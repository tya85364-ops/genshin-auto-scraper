from playwright.sync_api import sync_playwright
import re, time, random, requests, json, os, schedule
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials

BASE_URL = "https://www.8591.com.tw"
MAX_PAGES = 10
TRASH_KEYWORDS = ["徵", "代練", "初始號"]
BIG_SELLER_THRESHOLD = 5
RECENT_DAYS = 10
PRICE_DROP_THRESHOLD = 0.15
DISCORD_PRICE_DROP = "https://discord.com/api/webhooks/1485605514284105770/F0fPL4zY9MVr0xDq3yXFwVawG4d-mdsGB9D1H1TjdCcAjwPEdzbluMbUyqXmMC8sV-Tu"
BARGAIN_THRESHOLD = 0.50
PLATFORM_FEE = 0.94
SPREADSHEET_ID = "1SOt-2DwJVEcEgvuvQfAvW6ue6WcrnvywxPbKIJFEcYI"
GCP_KEY_FILE = "/app/gcp_key.json"
PRICE_TRACKER_FILE = "price_tracker.json"

TIER_LIST_FILES = {
 "原神": "C:\\Users\\toge\\genshin_tier_list.json",
 "鳴潮": "C:\\Users\\toge\\wutheringwaves_tier_list.json",
 "崩鐵": "C:\\Users\\toge\\hsr_tier_list.json",
}

TIER_WEIGHTS = {
 "tierSS": 10,
 "tierS": 7,
 "tierA": 4,
 "tierB": 2,
 "tierC": 1,
 "tierD": 1,
}

HIGH_TIER_LEVELS = {"tierSS", "tierS"}

HEADERS = ["發現時間", "上架時間", "標題", "價格", "金角", "金武/專",
 "純角CP", "含武CP", "加權CP", "預估獲利", "滿命角色", "優於均值", "賣家ID", "連結"]

COMPLETED_HEADERS = [
 "成交發現日", "上架時間", "售出所需天數",
 "標題", "價格", "金角", "金武/專",
 "滿命角色", "高Tier角色(SS/S)",
 "純角CP", "含武CP", "加權CP",
 "賣家ID", "連結"
]

# ===================== Tier List 載入 =====================

def load_tier_weights(game_name):
 filepath = TIER_LIST_FILES.get(game_name)
 if not filepath or not os.path.exists(filepath):
 print(f" ⚠️ 找不到 {game_name} Tier List，使用預設權重")
 return {}, [], {}, set()

 with open(filepath, "r", encoding="utf-8") as f:
 data = json.load(f)

 char_weights = {}
 new_chars = data.get("highValueFor8591", [])
 high_tier_chars = set()

 for tier, base_weight in TIER_WEIGHTS.items():
 for char in data.get(tier, []):
 name = char["name"]
 consensus = char.get("consensus", 1)
 weight = base_weight + min(consensus - 1, 3)
 char_weights[name] = weight
 if tier in HIGH_TIER_LEVELS:
 high_tier_chars.add(name)

 print(f" 📊 {game_name} Tier List 載入：{len(char_weights)} 個角色（SS/S級：{len(high_tier_chars)}個）")
 return char_weights, new_chars, {}, high_tier_chars

def build_games_config():
 base_alias = {
 "原神": {
 "水神": "芙寧娜", "芙芙": "芙寧娜",
 "草神": "納西妲", "小草神": "納西妲",
 "雷神": "雷電將軍", "影": "雷電將軍",
 "萬葉": "楓原萬葉", "葉天帝": "楓原萬葉",
 "水龍": "那維萊特", "水龍王": "那維萊特",
 "火神": "瑪薇卡",
 "僕人": "阿蕾奇諾", "小丑": "阿蕾奇諾",
 "綾華": "神里綾華", "達達": "達達利亞",
 "鍾離": "鐘離",
 },
 "鳴潮": {"卡卡": "卡卡羅"},
 "崩鐵": {"花火": "花火", "阮梅": "阮•梅"},
 }

 games = {
 "原神": {
 "emoji": "⚙️",
 "list_url": "https://www.8591.com.tw/v3/mall/list/34169?searchGame=34169&searchServer=34170&searchType=2&priceStart=100&priceEnd=200000&post_time_sort=1",
 "completed_url": "https://www.8591.com.tw/v3/mall/list/34169?searchGame=34169&searchServer=34170&searchType=2&priceStart=100&priceEnd=200000&completed=1&post_time_sort=1",
 "discord": "https://discord.com/api/webhooks/1483040957397205014/kyBaVtkZ4s0ECMBdNz0RT1eVG8A4irguVG-VzsNBL_TVTd5wTWU4eHHZHpq2cRnMjNXY",
 "discord_bargain": "https://discord.com/api/webhooks/1484537008935276624/VemiDQp698IAADQOFWUnLU9x5kDzlZ2NfwtqdSk9FdIlHcvFnyyAV_LNeMXm9qbgHfzr",
 "discord_maxconst": "https://discord.com/api/webhooks/1484537441405767773/5Jxr2h9BzgkgeVj6adxFA7_Eysx-cOI7PlI-YmfrdTq_ukKob21jLwIc9QRkIsHyGynV",
 "excel": "C:\\Users\\toge\\genshin_listings.xlsx",
 "stats_file": "gs_market_stats.json",
 "history_file": "gs_completed_history.json",
 "listing_seen_file": "gs_listing_seen.json",
 "seller_file": "global_sellers.json",
 },
 "鳴潮": {
 "emoji": "🌊",
 "list_url": "https://www.8591.com.tw/v3/mall/list/53396?searchGame=53396&searchServer=53397&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&post_time_sort=1",
 "completed_url": "https://www.8591.com.tw/v3/mall/list/53396?searchGame=53396&searchServer=53397&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&completed=1&post_time_sort=1",
 "discord": "https://discord.com/api/webhooks/1483459423434182798/iDMSDYDlZ5bp0_sMPHCFUISyYQxlkO5fzMJP9jo6NLXEHC_AGZkMN8Nb0SmoCuk-c2P9",
 "discord_bargain": "https://discord.com/api/webhooks/1484537092527620117/dCVNGjSHXuTj3MO24vlndbaHuXcqKpolgIMckjpyNsBurXYwEzuwCrIr3LpUI_C0ilUI",
 "discord_maxconst": "https://discord.com/api/webhooks/1484537497827541003/2ZMhGXZeBXq7vmwEUk_fZ6OgtRiUkShi0dQ2ZE2Z8M_XQLK6lrDp356offRRbJB4u94R",
 "excel": "C:\\Users\\toge\\wuwa_listings.xlsx",
 "stats_file": "ww_market_stats.json",
 "history_file": "ww_completed_history.json",
 "listing_seen_file": "ww_listing_seen.json",
 "seller_file": "global_sellers.json",
 },
 "崩鐵": {
 "emoji": "🚂",
 "list_url": "https://www.8591.com.tw/v3/mall/list/44693?searchGame=44693&searchServer=53160&searchType=2&accountTag=3&priceStart=100&priceEnd=100000&post_time_sort=1",
 "completed_url": "https://www.8591.com.tw/v3/mall/list/44693?searchGame=44693&searchServer=53160&searchType=2&accountTag=3&priceStart=100&priceEnd=100000&completed=1&post_time_sort=1",
 "discord": "https://discord.com/api/webhooks/1483469454376566943/QIaka_rST9Af8dQayNIKf11zM4a6X06k-3MKFIHbB0kK9AJpcR3Lp6fSys_xeA2oGlZT",
 "discord_bargain": "https://discord.com/api/webhooks/1484537146915291176/gHNGq_1m3j_jTJfZzGK2LSOKUxg1QAHmi_ejCP2Frtb9Qg0X1tkVBC4PihH34WObL0u1",
 "discord_maxconst": "https://discord.com/api/webhooks/1484537547689431222/36fxNxtjrLg2LAmvwbrKHE4EiWpor3uJD8mJxQUghpiZd3X1GncMHOJSgG0TH4V32LwP",
 "excel": "C:\\Users\\toge\\starrail_listings.xlsx",
 "stats_file": "sr_market_stats.json",
 "history_file": "sr_completed_history.json",
 "listing_seen_file": "sr_listing_seen.json",
 "seller_file": "global_sellers.json",
 },
 "絕區零": {
 "emoji": "📼",
 "list_url": "https://www.8591.com.tw/v3/mall/list/49016?searchGame=49016&searchServer=57343&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&post_time_sort=1",
 "completed_url": "https://www.8591.com.tw/v3/mall/list/49016?searchGame=49016&searchServer=57343&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&completed=1&post_time_sort=1",
 "discord": "https://discord.com/api/webhooks/1491093750057795838/WoxqQbxWapT6NCdYannyMY-JYEvtojfwoHz2HO_1CyymsXOnVvUuE9tNeLjvcZriDPTH",
 "discord_bargain": "https://discord.com/api/webhooks/1491093789975122010/y73z2n2-4NA4Zsshg9W9Wsvz11ko-gxOBpWBT_u_YWdguj-HW28-aMhhU3b6Z7fsz72t",
 "discord_maxconst": "https://discord.com/api/webhooks/1491094260274167828/Oe1rX1zf9IDdN4a_gTl2eHmxHSnSS0gjIlZ2qWmof4vPA7eHGckHpsQW7t7umo0meeLp",
 "excel": "C:\\Users\\toge\\zzz_listings.xlsx",
 "stats_file": "zzz_market_stats.json",
 "history_file": "zzz_completed_history.json",
 "listing_seen_file": "zzz_listing_seen.json",
 "seller_file": "global_sellers.json",
 },
 }

 for game_name, g in games.items():
 char_weights, new_chars, _, high_tier_chars = load_tier_weights(game_name)
 g["char_weights"] = char_weights
 g["new_chars"] = new_chars
 g["alias_map"] = base_alias.get(game_name, {})
 g["high_tier_chars"] = high_tier_chars

 return games

# ===================== listing_seen 新版（含時間戳記）=====================

def load_listing_seen(filepath):
    _id = os.path.basename(filepath)
    col = _mongo("listing_seen")
    if col is not None:
        try:
            doc = col.find_one({"_id": _id})
            if doc:
                seen_map = doc.get("seen_map", {})
                for url in doc.get("urls", []):
                    if url not in seen_map:
                        seen_map[url] = {"first_seen": ""}
                    elif isinstance(seen_map[url], str):
                        seen_map[url] = {"first_seen": seen_map[url]}
                return seen_map
        except Exception as e:
            print(f"  MongoDB load_listing_seen 失敗：{e}")
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        urls = data.get("urls", [])
        seen_map = data.get("seen_map", {})
        for url in urls:
            if url not in seen_map:
                seen_map[url] = {"first_seen": ""}
            elif isinstance(seen_map[url], str):
                seen_map[url] = {"first_seen": seen_map[url]}
        return seen_map
    return {}

def save_listing_seen(filepath, seen_map):
    data = {
        "urls": list(seen_map.keys()),
        "seen_map": seen_map,
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    col = _mongo("listing_seen")
    if col is not None:
        try:
            _id = os.path.basename(filepath)
            col.replace_one({"_id": _id}, {**data, "_id": _id}, upsert=True)
        except Exception as e:
            print(f"  MongoDB save_listing_seen 失敗：{e}")

def _first_seen_date(entry):
 """seen_map entry (str or dict) -> first_seen string"""
 if isinstance(entry, dict):
  return entry.get("first_seen", "")
 return entry or ""

def find_listing_date_by_seller(seller_id, title, seen_map):
 """
 URL 找不到時，用 seller_id 找同賣家最早 first_seen。
 title 有提供時，優先選金角數相符的那筆。
 """
 if not seller_id:
  return ""
 import re as _re
 gm = _re.search(r'(\d+)\u91d1\u89d2', title) if title else None
 gold_char = int(gm.group(1)) if gm else -1
 best_date, best_score = "", -1
 for entry in seen_map.values():
  if not isinstance(entry, dict):
   continue
  if entry.get("seller_id", "") != seller_id:
   continue
  first_seen = entry.get("first_seen", "")
  if not first_seen:
   continue
  stored = entry.get("title", "")
  score = 0
  if title and stored:
   score += sum(1 for w in ["\u91d1\u89d2", "\u91d1\u6b66", "\u91d1\u5c08"] if w in title and w in stored)
   if gold_char >= 0:
    gm2 = _re.search(r'(\d+)\u91d1\u89d2', stored)
    if gm2 and int(gm2.group(1)) == gold_char:
     score += 3
  if score > best_score:
   best_score, best_date = score, first_seen
 return best_date

def calc_days_on_market(post_time_str, seen_map, url):
 """
 計算售出所需天數：
 優先用詳情頁抓到的上架時間，
 其次用 listing_seen 的首次發現時間，
 都沒有就返回 -
 """
 today = datetime.now()

 # 方法A：用詳情頁的上架時間
 if post_time_str and post_time_str != "-":
 for fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
 try:
 d = datetime.strptime(post_time_str.strip(), fmt)
 days = (today - d).days
 return str(days)
 except:
 continue

 # 方法B：用 listing_seen 首次發現時間（支援 str/dict）
 first_seen = _first_seen_date(seen_map.get(url, ""))
 if first_seen:
  try:
   d = datetime.strptime(first_seen, "%Y-%m-%d")
   days = (today - d).days
   return f"≥{days}" # 加 ≥ 表示這是最少天數，實際可能更長
  except:
   pass

 return "-"

# ===================== Google Sheets =====================

def get_gsheet():
 scopes = ["https://spreadsheets.google.com/feeds",
 "https://www.googleapis.com/auth/drive"]
 creds = Credentials.from_service_account_file(GCP_KEY_FILE, scopes=scopes)
 return gspread.authorize(creds)

def init_gsheet(gc, game_name):
 try:
 sh = gc.open_by_key(SPREADSHEET_ID)
 try:
 ws = sh.worksheet(game_name)
 except:
 ws = sh.add_worksheet(title=game_name, rows=5000, cols=14)
 if not ws.get_all_values() or ws.cell(1, 1).value != "發現時間":
 ws.insert_row(HEADERS, 1)
 if ws.row_count < 5000 or ws.col_count < 14:
 ws.resize(rows=5000, cols=14)
 return ws
 except Exception as e:
 print(f" Google Sheets 初始化失敗：{e}")
 return None

def init_gsheet_completed(gc, game_name):
 sheet_name = f"{game_name}-成交紀錄"
 try:
 sh = gc.open_by_key(SPREADSHEET_ID)
 try:
 ws = sh.worksheet(sheet_name)
 except:
 ws = sh.add_worksheet(title=sheet_name, rows=5000, cols=14)
 if not ws.get_all_values() or ws.cell(1, 1).value != "成交發現日":
 ws.insert_row(COMPLETED_HEADERS, 1)
 if ws.row_count < 5000 or ws.col_count < 14:
 ws.resize(rows=5000, cols=14)
 return ws
 except Exception as e:
 print(f" 成交紀錄分頁初始化失敗：{e}")
 return None

def gsheet_batch_insert(ws, rows_to_add):
 if not rows_to_add:
 return
 for i in range(0, len(rows_to_add), 10):
 batch = rows_to_add[i:i+10]
 for row in batch:
 ws.insert_row(row, 2)
 time.sleep(2)
 if i + 10 < len(rows_to_add):
 print(f" 已寫入 {i+len(batch)}/{len(rows_to_add)}...")
 time.sleep(10)

def update_gsheet(ws, new_items, thresholds, sellers):
    if not ws:
        return
    try:
        # 讀取所有列：url -> (row_idx_1based, min_str, max_str)
        all_values = ws.get_all_values()
        url_row_map = {}
        for row_idx, row in enumerate(all_values[1:], start=2):
            url = row[13] if len(row) > 13 else ""
            if url:
                col_o = row[14].strip() if len(row) > 14 else ""
                col_p = row[15].strip() if len(row) > 15 else ""
                url_row_map[url] = (row_idx, col_o, col_p)

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        rows_to_add = []
        min_max_updates = []  # (row_idx, new_o, new_p)

        for r in new_items:
            url = r['url']
            price = r['price']

            if url in url_row_map:
                # 已存在 -> 只比對更新 O/P 欄（歷史低/高價）
                row_idx, col_o, col_p = url_row_map[url]
                new_o, new_p = col_o, col_p
                changed = False

                if not new_o or new_o == "-":
                    new_o = str(price)
                    changed = True
                else:
                    try:
                        if price < float(new_o.replace(',', '')):
                            new_o = str(price)
                            changed = True
                    except ValueError:
                        new_o = str(price)
                        changed = True

                if not new_p or new_p == "-":
                    new_p = str(price)
                    changed = True
                else:
                    try:
                        if price > float(new_p.replace(',', '')):
                            new_p = str(price)
                            changed = True
                    except ValueError:
                        new_p = str(price)
                        changed = True

                if changed:
                    min_max_updates.append((row_idx, new_o, new_p))
                continue

            # 全新列 -> 準備插入
            cp1, cp2, cpw = r['cp1'], r['cp2'], r['cpw']
            is_good = (cp1 <= thresholds['cp1_threshold'] or
                       cp2 <= thresholds['cp2_threshold'] or
                       cpw <= thresholds['cpw_threshold'])
            good_str = "✅ 優於均值" if is_good else ""
            const_str = ", ".join(r.get('max_const', []))
            cp1_str = f"{cp1:.2f}" if cp1 != float('inf') else "-"
            cp2_str = f"{cp2:.2f}" if cp2 != float('inf') else "-"
            cpw_str = f"{cpw:.2f}" if cpw != float('inf') else "-"
            profit = r.get('estimated_profit')
            profit_str = f"+${profit:,.0f}" if profit and profit > 0 else (f"-${abs(profit):,.0f}" if profit else "-")
            sid = r.get('seller_id', '')
            is_big = sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD
            seller_str = f"🍽️{sid}" if is_big else sid
            rows_to_add.append([
                now_str, r.get('post_time', ''), r['title'], price,
                r['gold_char'], r['gold_weap'], cp1_str, cp2_str, cpw_str,
                profit_str, const_str, good_str, seller_str, url,
                str(price), str(price)
            ])
            url_row_map[url] = (-1, str(price), str(price))

        # 批次更新 O/P 欄
        if min_max_updates:
            batch = [{'range': f'O{idx}:P{idx}', 'values': [[o, p]]}
                     for idx, o, p in min_max_updates]
            for i in range(0, len(batch), 500):
                ws.batch_update(batch[i:i+500])
                time.sleep(1)
            print(f"  min/max 更新：{len(min_max_updates)} 筆")

        if not rows_to_add:
            print("  Google Sheets：無新資料")
            return
        gsheet_batch_insert(ws, rows_to_add)
        print(f"  Google Sheets 更新：新增 {len(rows_to_add)} 筆")
    except Exception as e:
        print(f"  Google Sheets 更新失敗：{e}")

def update_gsheet_completed(ws, new_trades, sellers, seen_map, high_tier_chars):
 if not ws or not new_trades:
 return
 try:
 existing = ws.col_values(14)
 existing_urls = set(existing[1:])
 now_str = datetime.now().strftime("%Y-%m-%d")
 rows_to_add = []

 for r in new_trades:
 if r['url'] in existing_urls:
 continue
 cp1, cp2, cpw = r['cp1'], r['cp2'], r['cpw']
 cp1_str = f"{cp1:.2f}" if cp1 != float('inf') else "-"
 cp2_str = f"{cp2:.2f}" if cp2 != float('inf') else "-"
 cpw_str = f"{cpw:.2f}" if cpw != float('inf') else "-"

 # 上架時間（詳情頁抓到的）
 post_time = r.get('post_time', '')
 # 過濾掉括號數字（列表頁誤讀的成交數）
 if post_time and re.match(r'^\(\d+\)$', post_time.strip()):
 post_time = ''

  # Fallback 1: URL 直查 listing_seen
  if not post_time:
   post_time = _first_seen_date(seen_map.get(r['url'], ''))
  # Fallback 2: 同賣家 fuzzy match（改價後 URL 換掉也能補）
  if not post_time:
   post_time = find_listing_date_by_seller(r.get('seller_id', ''), r.get('title', ''), seen_map)

  # 售出所需天數（優先詳情頁時間，其次 listing_seen）
  days = calc_days_on_market(post_time, seen_map, r['url'])

 const_str = ", ".join(r.get('max_const', []))

 # 高Tier角色
 title = r.get('title', '')
 high_tier_found = [char for char in high_tier_chars if char in title]
 high_tier_str = ", ".join(high_tier_found) if high_tier_found else "-"

 # 賣家ID
 sid = r.get('seller_id', '')
 is_big = sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD
 seller_str = f"🍽️{sid}（大盤商）" if is_big else sid

 rows_to_add.append([
 now_str,
 post_time if post_time else "-",
 days,
 r['title'],
 r['price'],
 r['gold_char'],
 r['gold_weap'],
 const_str if const_str else "-",
 high_tier_str,
 cp1_str,
 cp2_str,
 cpw_str,
 seller_str,
 r['url'],
 ])
 existing_urls.add(r['url'])

 if not rows_to_add:
 print(" 成交紀錄：無新資料")
 return
 gsheet_batch_insert(ws, rows_to_add)
 print(f" 成交紀錄更新：新增 {len(rows_to_add)} 筆")
 except Exception as e:
 print(f" 成交紀錄更新失敗：{e}")

# ===================== 降價追蹤 =====================

_mongo_client = None

def _mongo(collection):
    """取得指定 MongoDB collection，無 MONGODB_URI 或失敗時回傳 None。"""
    global _mongo_client
    uri = os.getenv("MONGODB_URI", "")
    if not uri:
        return None
    try:
        from pymongo import MongoClient
        if _mongo_client is None:
            _mongo_client = MongoClient(uri, serverSelectionTimeoutMS=5000)
        # URI 有帶 db name 就用 get_default_database，否則 fallback scraper_db
        try:
            db = _mongo_client.get_default_database()
        except Exception:
            db = _mongo_client["scraper_db"]
        return db[collection]
    except Exception as e:
        print(f"  MongoDB 連線失敗（本地模式）：{e}")
        return None

def load_price_tracker():
    """優先從 MongoDB 讀，失敗則讀本地 JSON。"""
    col = _mongo("price_tracker")
    if col is not None:
        try:
            doc = col.find_one({"_id": "price_tracker"})
            if doc:
                doc.pop("_id", None)
                print(f"  price_tracker 從 MongoDB 載入（共 {sum(len(v) for v in doc.values())} 筆）")
                return doc
        except Exception as e:
            print(f"  MongoDB 讀取失敗：{e}")
    # Fallback: 本地 JSON
    if os.path.exists(PRICE_TRACKER_FILE):
        with open(PRICE_TRACKER_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        print(f"  price_tracker 從本地 JSON 載入（共 {sum(len(v) for v in data.values())} 筆）")
        return data
    return {}

def save_price_tracker(tracker):
    """同時寫入 MongoDB 與本地 JSON（互為備援）。"""
    # 寫本地 JSON（本機執行 or 備份用）
    try:
        with open(PRICE_TRACKER_FILE, "w", encoding="utf-8") as f:
            json.dump(tracker, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  price_tracker 本地儲存失敗：{e}")

    # 寫 MongoDB
    col = _mongo("price_tracker")
    if col is not None:
        try:
            doc = {**tracker, "_id": "price_tracker"}
            col.replace_one({"_id": "price_tracker"}, doc, upsert=True)
            total = sum(len(v) for v in tracker.values())
            print(f"  price_tracker 已存 MongoDB（共 {total} 筆）")
        except Exception as e:
            print(f"  MongoDB 儲存失敗：{e}")

def check_price_drop(tracker, listings, discord_url, emoji, name):
 dropped = []
 for r in listings:
  url, price = r['url'], r['price']
  update_tracker = True
  if url in tracker:
   old_price = tracker[url]['price']
   if old_price > price:
    drop_pct = (old_price - price) / old_price
    if drop_pct >= PRICE_DROP_THRESHOLD:
     dropped.append({**r, 'old_price': old_price, 'drop_pct': drop_pct})
     update_tracker = True
    else:
     update_tracker = False
   elif price > old_price:
    update_tracker = True
   else:
    update_tracker = False
  else:
   update_tracker = True
  if update_tracker:
   tracker[url] = {'price': price, 'updated': datetime.now().strftime("%Y-%m-%d %H:%M")}
 if dropped:
  msg = f"{emoji} **🔥 {name} 降價警告！（共{len(dropped)}個）**\n{'─'*38}\n"
  for r in dropped:
   line = (
    f"**降價 {r['drop_pct']*100:.0f}%！** ~~${r['old_price']:,}~~ → **${r['price']:,}**\n"
    f"{r['title']}\n純角{r['cp1']:.1f} 含武{r['cp2']:.1f}\n{r['url']}\n\n"
   )
   if len(msg) + len(line) > 1900:
    send_discord(discord_url, msg)
    msg = f"{emoji} **🔥 降價警告（續）**\n"
   msg += line
  if msg.strip():
   send_discord(discord_url, msg)
 return tracker

# ===================== 獲利預估 =====================

def estimate_profit(r, stats):
 if not stats.get("records"):
 return None
 price, gold_char = r['price'], r['gold_char']
 if gold_char <= 0:
 return None
 similar = [rec for rec in stats["records"]
 if abs(rec.get("gold_char", 0) - gold_char) <= 3 and rec.get("price", 0) > 0]
 if len(similar) < 3:
 if stats["count"] > 0:
 avg_price = stats["price_sum"] / stats["count"]
 ratio = gold_char / max(stats["gold_char_sum"] / stats["count"], 1)
 estimated_resale = avg_price * ratio
 else:
 return None
 else:
 estimated_resale = sum(rec["price"] for rec in similar) / len(similar)
 return round((estimated_resale * PLATFORM_FEE) - price)

# ===================== 核心解析 =====================

def resolve_alias(name, alias_map):
 return alias_map.get(name, name)

def parse_title_smart(title, char_weights, alias_map):
 gold_char, gold_weap, weighted_score = 0, 0, 0
 max_const_chars = []

 m_char = re.search(r'(\d+)金角', title)
 m_weap = re.search(r'(\d+)金(?:武|專)', title)
 if m_char:
 gold_char = int(m_char.group(1))
 if m_weap:
 gold_weap = int(m_weap.group(1))

 plus_patterns = re.findall(r'(\d+)\+(\d+)([\u4e00-\u9fff]{1,4})', title)
 for n1, n2, char in plus_patterns:
 char = resolve_alias(char, alias_map)
 n1, n2 = int(n1), int(n2)
 if gold_char == 0:
 gold_char += n1
 if gold_weap == 0:
 gold_weap += n2
 weight = char_weights.get(char, 1)
 weighted_score += n1 * weight * 10 + n2 * weight * 5
 if n1 >= 6 and char not in max_const_chars:
 max_const_chars.append(char)

 const_patterns = re.findall(r'(\d+)命([\u4e00-\u9fff]{1,4})', title)
 const_patterns += [(b, a) for a, b in re.findall(r'([\u4e00-\u9fff]{1,4})\s*(\d+)命', title)]
 for n, char in const_patterns:
 char = resolve_alias(char, alias_map)
 n = int(n)
 weight = char_weights.get(char, 1)
 weighted_score += (n + 1) * weight * 10
 if n >= 6 and char not in max_const_chars:
 max_const_chars.append(char)

 for pattern in [
 r'(?:滿命|C6|E6)\s*([\u4e00-\u9fff]{2,4})',
 r'([\u4e00-\u9fff]{2,4})\s*(?:滿命|C6|E6)',
 r'([\u4e00-\u9fff]{2,4})\s*6命',
 r'6命\s*([\u4e00-\u9fff]{2,4})',
 ]:
 for c in re.findall(pattern, title):
 c = resolve_alias(c, alias_map)
 if len(c) >= 2 and c not in max_const_chars:
 max_const_chars.append(c)

 if not plus_patterns and not const_patterns:
 for char, weight in char_weights.items():
 aliases = [k for k, v in alias_map.items() if v == char]
 if char in title or any(a in title for a in aliases):
 weighted_score += weight * 10

 return gold_char, gold_weap, weighted_score, max_const_chars

def parse_detail_for_gold(page, url, title):
 try:
 page.goto(url, timeout=30000)
 time.sleep(random.uniform(2, 3))
 body_text = page.inner_text("body")
 gold_char, gold_weap = 0, 0
 m_char = re.search(r'(\d+)金角', body_text)
 m_weap = re.search(r'(\d+)金(?:武|專)', body_text)
 if m_char:
 gold_char = int(m_char.group(1))
 if m_weap:
 gold_weap = int(m_weap.group(1))
 # 抓上架時間（格式：YYYY-MM-DD 或 YYYY/MM/DD）
 time_matches = re.findall(r'(\d{4}[-/]\d{2}[-/]\d{2})', body_text)
 post_time = time_matches[0] if time_matches else ""
 return gold_char, gold_weap, post_time
 except:
 return 0, 0, ""

def cp_char_only(price, gold_char):
 if price <= 0 or gold_char <= 0:
 return float('inf')
 return price / (gold_char * 10)

def cp_with_weap(price, gold_char, gold_weap):
 total = gold_char * 10 + gold_weap * 5
 if price <= 0 or total <= 0:
 return float('inf')
 return price / total

def cp_weighted(price, weighted_score):
 if price <= 0 or weighted_score <= 0:
 return float('inf')
 return price / weighted_score

def is_recent(post_time_str, days=RECENT_DAYS):
 if not post_time_str:
 return False
 if "分鐘前" in post_time_str or "小時前" in post_time_str:
 return True
 if "天前" in post_time_str:
 m = re.search(r'(\d+)天前', post_time_str)
 if m and int(m.group(1)) <= days:
 return True
 try:
 for fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
 try:
 d = datetime.strptime(post_time_str.strip(), fmt)
 return (datetime.now() - d).days <= days
 except:
 continue
 except:
 pass
 return False

def send_discord(webhook_url, content):
 webhook_url = webhook_url.strip().rstrip('ㄛ')
 try:
 r = requests.post(webhook_url, json={"content": content})
 print(f" Discord：{r.status_code}")
 time.sleep(0.8)
 except Exception as e:
 print(f" Discord 失敗：{e}")

def load_stats(filepath):
    _id = os.path.basename(filepath)
    col = _mongo("stats")
    if col is not None:
        try:
            doc = col.find_one({"_id": _id})
            if doc:
                doc.pop("_id", None)
                if "records" not in doc:
                    doc["records"] = []
                return doc
        except Exception as e:
            print(f"  MongoDB load_stats 失敗：{e}")
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        if "records" not in data:
            data["records"] = []
        return data
    return {"count": 0, "cp1_sum": 0, "cp2_sum": 0, "cpw_sum": 0,
            "price_sum": 0, "gold_char_sum": 0, "records": [], "last_updated": ""}

def update_stats(stats, new_trades, filepath):
 valid = [r for r in new_trades if r['cp1'] != float('inf')]
 if "records" not in stats:
 stats["records"] = []
 for r in valid:
 stats["count"] += 1
 stats["cp1_sum"] += r['cp1']
 stats["cp2_sum"] += r['cp2'] if r['cp2'] != float('inf') else r['cp1']
 stats["cpw_sum"] += r['cpw'] if r['cpw'] != float('inf') else r['cp1']
 stats["price_sum"] += r['price']
 stats["gold_char_sum"] += r['gold_char']
 stats["records"].append({
 "date": datetime.now().strftime("%Y-%m-%d"),
 "price": r['price'],
 "gold_char": r['gold_char'],
 "gold_weap": r['gold_weap'],
 "cp1": round(r['cp1'], 2),
 "cp2": round(r['cp2'], 2) if r['cp2'] != float('inf') else None,
 "cpw": round(r['cpw'], 2) if r['cpw'] != float('inf') else None,
 })
 stats["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
 try:
  with open(filepath, "w", encoding="utf-8") as f:
   json.dump(stats, f, ensure_ascii=False, indent=2)
 except Exception:
  pass
 col = _mongo("stats")
 if col is not None:
  try:
   _id = os.path.basename(filepath)
   col.replace_one({"_id": _id}, {**stats, "_id": _id}, upsert=True)
  except Exception as e:
   print(f"  MongoDB save_stats 失敗：{e}")
 return stats

def get_thresholds(stats):
 if stats["count"] == 0:
 return {"cp1_threshold": 30, "cp2_threshold": 25, "cpw_threshold": 20,
 "cp1_avg": 30, "cp2_avg": 25, "cpw_avg": 20,
 "price_avg": 0, "gold_char_avg": 0}
 n = stats["count"]
 cp1_avg = stats["cp1_sum"] / n
 cp2_avg = stats["cp2_sum"] / n
 cpw_avg = stats["cpw_sum"] / n
 return {
 "cp1_avg": cp1_avg, "cp2_avg": cp2_avg, "cpw_avg": cpw_avg,
 "cp1_threshold": cp1_avg * 0.8,
 "cp2_threshold": cp2_avg * 0.8,
 "cpw_threshold": cpw_avg * 0.8,
 "price_avg": stats["price_sum"] / n,
 "gold_char_avg": stats["gold_char_sum"] / n,
 }

def load_seen(filepath, key="urls"):
    _id = os.path.basename(filepath)
    col = _mongo("seen")
    if col is not None:
        try:
            doc = col.find_one({"_id": _id})
            if doc:
                return set(doc.get(key, []))
        except Exception as e:
            print(f"  MongoDB load_seen 失敗：{e}")
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return set(json.load(f).get(key, []))
    return set()

def save_seen(filepath, seen, key="urls"):
    data = {key: list(seen), "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")}
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass
    col = _mongo("seen")
    if col is not None:
        try:
            _id = os.path.basename(filepath)
            col.replace_one({"_id": _id}, {**data, "_id": _id}, upsert=True)
        except Exception as e:
            print(f"  MongoDB save_seen 失敗：{e}")

def load_sellers(filepath):
    _id = os.path.basename(filepath)
    col = _mongo("sellers")
    if col is not None:
        try:
            doc = col.find_one({"_id": _id})
            if doc:
                doc.pop("_id", None)
                return doc
        except Exception as e:
            print(f"  MongoDB load_sellers 失敗：{e}")
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def update_sellers(sellers, trades, filepath):
 for r in trades:
 sid = r.get("seller_id", "")
 if not sid:
 continue
 if sid not in sellers:
 sellers[sid] = {"count": 0, "prices": [], "titles": []}
 sellers[sid]["count"] += 1
 sellers[sid]["prices"].append(r["price"])
 sellers[sid]["titles"].append(r["title"][:30])
 sellers[sid]["prices"] = sellers[sid]["prices"][-50:]
 sellers[sid]["titles"] = sellers[sid]["titles"][-50:]
 try:
  with open(filepath, "w", encoding="utf-8") as f:
   json.dump(sellers, f, ensure_ascii=False, indent=2)
 except Exception:
  pass
 col = _mongo("sellers")
 if col is not None:
  try:
   _id = os.path.basename(filepath)
   col.replace_one({"_id": _id}, {**sellers, "_id": _id}, upsert=True)
  except Exception as e:
   print(f"  MongoDB update_sellers 失敗：{e}")
 return sellers

def init_excel(filepath):
 wb = Workbook()
 ws = wb.active
 ws.title = "賣場列表"
 header_fill = PatternFill("solid", fgColor="1F4E79")
 header_font = Font(bold=True, color="FFFFFF", size=11)
 for col, h in enumerate(HEADERS, 1):
 cell = ws.cell(row=1, column=col, value=h)
 cell.fill = header_fill
 cell.font = header_font
 cell.alignment = Alignment(horizontal="center", vertical="center")
 col_widths = [16, 12, 60, 10, 8, 8, 10, 10, 10, 12, 20, 10, 12, 45]
 for i, w in enumerate(col_widths, 1):
 ws.column_dimensions[get_column_letter(i)].width = w
 ws.freeze_panes = "A2"
 wb.save(filepath)

def update_excel(filepath, new_items, thresholds, sellers):
 if not os.path.exists(filepath):
 init_excel(filepath)
 wb = load_workbook(filepath)
 ws = wb.active
 existing_urls = set()
 for row in ws.iter_rows(min_row=2, values_only=True):
 if row[13]:
 existing_urls.add(row[13])
 now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
 added = 0
 for r in new_items:
 if r['url'] in existing_urls:
 continue
 cp1, cp2, cpw = r['cp1'], r['cp2'], r['cpw']
 is_good = (cp1 <= thresholds['cp1_threshold'] or
 cp2 <= thresholds['cp2_threshold'] or
 cpw <= thresholds['cpw_threshold'])
 good_str = "✅ 優於均值" if is_good else ""
 const_str = ", ".join(r.get('max_const', []))
 cp1_str = f"{cp1:.2f}" if cp1 != float('inf') else "-"
 cp2_str = f"{cp2:.2f}" if cp2 != float('inf') else "-"
 cpw_str = f"{cpw:.2f}" if cpw != float('inf') else "-"
 profit = r.get('estimated_profit')
 profit_str = f"+${profit:,.0f}" if profit and profit > 0 else (f"-${abs(profit):,.0f}" if profit else "-")
 sid = r.get('seller_id', '')
 is_big = sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD
 seller_str = f"🍽️{sid}" if is_big else sid
 ws.insert_rows(2)
 row_data = [now_str, r.get('post_time', ''), r['title'], r['price'],
 r['gold_char'], r['gold_weap'], cp1_str, cp2_str, cpw_str,
 profit_str, const_str, good_str, seller_str, r['url']]
 for col, val in enumerate(row_data, 1):
 cell = ws.cell(row=2, column=col, value=val)
 cell.alignment = Alignment(vertical="center", wrap_text=True)
 if is_good and const_str:
 cell.fill = PatternFill("solid", fgColor="FFC000")
 elif is_good:
 cell.fill = PatternFill("solid", fgColor="E2EFDA")
 elif const_str:
 cell.fill = PatternFill("solid", fgColor="FFE699")
 elif is_big:
 cell.fill = PatternFill("solid", fgColor="FCE4D6")
 ws.row_dimensions[2].height = 40
 existing_urls.add(r['url'])
 added += 1
 wb.save(filepath)
 print(f" Excel 更新：新增 {added} 筆 → {filepath}")

def get_item_url(item, title_el):
 href = title_el.get_attribute("href")
 if href and href.startswith("/"):
 return BASE_URL + href
 elif href:
 return href
 else:
 im_el = item.query_selector("a[href^='im://']")
 if im_el:
 im_href = im_el.get_attribute("href") or ""
 m = re.search(r'i=(\d+)', im_href)
 if m:
 return f"{BASE_URL}/v3/mall/detail/{m.group(1)}"
 return ""

def scrape_pages(main_page, base_url, max_pages, label="",
 stop_at_seen=None, do_detail=False, detail_page=None,
 char_weights=None, alias_map=None):
 results = []
 seen_in_run = set()
 stop_at_seen = set(stop_at_seen) if stop_at_seen else set()
 hit_old = False

 for page_num in range(1, max_pages + 1):
 if hit_old:
 break
 first_row = (page_num - 1) * 40
 url = f"{base_url}&firstRow={first_row}"
 print(f" [{label}] 第{page_num}頁...")
 try:
 main_page.goto(url, timeout=60000)
 time.sleep(random.uniform(3, 4))
 items = main_page.query_selector_all("div.list-item")
 if not items:
 print(" 無商品，停止。")
 break
 print(f" 找到 {len(items)} 個")
 for item in items:
 try:
 title_el = item.query_selector("a.show-title")
 if not title_el:
 title_el = item.query_selector("span.show-title")
 price_el = item.query_selector("span.orange")
 if not price_el:
 price_el = item.query_selector("div.list-item-price")
 time_el = item.query_selector(".list-item-bread span.ml15")
 if not time_el:
 time_el = item.query_selector(".fc3")
 seller_el = item.query_selector("a[href^='im://']")
 seller_id = seller_el.get_attribute("data-fuid") if seller_el else ""

 if not title_el or not price_el:
 continue
 title = title_el.inner_text().strip()
 detail_url = get_item_url(item, title_el)
 price_text = re.sub(r'[^\d]', '', price_el.inner_text())
 if not price_text:
 continue
 price = int(price_text)
 post_time = time_el.inner_text().strip() if time_el else ""

 # 過濾列表頁誤讀的成交數（括號數字）
 if post_time and re.match(r'^\(\d+\)$', post_time.strip()):
 post_time = ''

 if any(kw in title for kw in TRASH_KEYWORDS):
 continue
 if price < 100:
 continue
 if not detail_url:
 continue
 if detail_url in seen_in_run:
 continue
 if detail_url in stop_at_seen:
 print(" 遇到已記錄交易，停止。")
 hit_old = True
 break

 seen_in_run.add(detail_url)
 gold_char, gold_weap, weighted, max_const = parse_title_smart(
 title, char_weights, alias_map)

 if do_detail and detail_page:
 d_char, d_weap, d_time = parse_detail_for_gold(detail_page, detail_url, title)
 if d_char and d_char > gold_char:
 gold_char = d_char
 if d_weap and d_weap > gold_weap:
 gold_weap = d_weap
 if d_time:
 post_time = d_time # 詳情頁時間優先

 results.append({
 "title": title, "price": price,
 "gold_char": gold_char, "gold_weap": gold_weap,
 "weighted": weighted,
 "cp1": cp_char_only(price, gold_char),
 "cp2": cp_with_weap(price, gold_char, gold_weap),
 "cpw": cp_weighted(price, weighted),
 "max_const": max_const,
 "post_time": post_time,
 "seller_id": seller_id,
 "url": detail_url
 })
 except Exception as e:
 continue
 except Exception as e:
 print(f" 頁面錯誤：{e}")
 break
 time.sleep(random.uniform(2, 3))
 return results

def format_item(r, cp_type="cp1", sellers=None):
 cp_val = r.get(cp_type, float('inf'))
 cp_label = {"cp1": "純角CP", "cp2": "含武CP", "cpw": "加權CP"}[cp_type]
 post_time = r.get('post_time', '')
 time_str = f" | 🆕 **{post_time}**" if is_recent(post_time) else (f" | 🕐 {post_time}" if post_time else "")
 const_str = f"\n⭐ **滿命：{', '.join(r['max_const'])}**" if r.get('max_const') else ""
 profit = r.get('estimated_profit')
 profit_str = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
 seller_id = r.get('seller_id', '')
 if seller_id and sellers and sellers.get(seller_id, {}).get("count", 0) >= BIG_SELLER_THRESHOLD:
 cnt = sellers[seller_id]["count"]
 avg_p = sum(sellers[seller_id]["prices"]) / len(sellers[seller_id]["prices"])
 seller_tag = f" | 🍽️ **大盤商 No.{seller_id}**（售{cnt}次 均${avg_p:,.0f}）"
 elif seller_id:
 seller_tag = f" | 賣家：{seller_id}"
 else:
 seller_tag = ""
 return (
 f"**標題：** {r['title']}\n"
 f"**價格：** ${r['price']:,} | 金角：{r['gold_char']} | 金武/專：{r['gold_weap']}{time_str}{seller_tag}{const_str}{profit_str}\n"
 f"**{cp_label}：** {cp_val:.2f}\n"
 f"**連結：** {r['url']}"
 )

def run_game(main_page, detail_page, game_key, g, gc, price_tracker):
 emoji = g["emoji"]
 name = game_key
 new_chars = g.get("new_chars", [])
 high_tier_chars = g.get("high_tier_chars", set())

 print(f"\n{'='*60}")
 print(f"{emoji} {name} | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
 print(f" 角色權重數：{len(g['char_weights'])} 個（SS/S級：{len(high_tier_chars)}個）")
 print(f"{'='*60}")

 stats = load_stats(g["stats_file"])
 seen_urls = load_seen(g["history_file"], key="seen_urls")
 listing_seen_map = load_listing_seen(g["listing_seen_file"])
 sellers = load_sellers(g["seller_file"])

 print(f"\n📊 [{name}] 抓取新成交紀錄...")
 new_completed = scrape_pages(
 main_page, g["completed_url"], 50, "已完成",
 stop_at_seen=seen_urls,
 do_detail=True, detail_page=detail_page,
 char_weights=g["char_weights"], alias_map=g["alias_map"])
 print(f" 新增 {len(new_completed)} 筆")

 if new_completed:
 stats = update_stats(stats, new_completed, g["stats_file"])
 sellers = update_sellers(sellers, new_completed, g["seller_file"])
 for r in new_completed:
 seen_urls.add(r["url"])
 save_seen(g["history_file"], seen_urls, key="seen_urls")

 try:
 ws_completed = init_gsheet_completed(gc, name)
 update_gsheet_completed(ws_completed, new_completed, sellers,
 listing_seen_map, high_tier_chars)
 except Exception as e:
 print(f" 成交紀錄寫入失敗：{e}")

 thresholds = get_thresholds(stats)
 print(f" 純角CP：{thresholds['cp1_avg']:.2f} → 門檻 {thresholds['cp1_threshold']:.2f}")
 print(f" 含武CP：{thresholds['cp2_avg']:.2f} → 門檻 {thresholds['cp2_threshold']:.2f}")
 print(f" 加權CP：{thresholds['cpw_avg']:.2f} → 門檻 {thresholds['cpw_threshold']:.2f}")

 big_sellers = {sid: info for sid, info in sellers.items() if info["count"] >= BIG_SELLER_THRESHOLD}
 seller_summary = ""
 if big_sellers:
 top = sorted(big_sellers.items(), key=lambda x: x[1]["count"], reverse=True)[:5]
 seller_summary = "\n🍽️ **常見大盤商：**\n"
 for sid, info in top:
 avg_price = sum(info["prices"]) / len(info["prices"]) if info["prices"] else 0
 seller_summary += f"No.{sid} | 出售{info['count']}次 | 均價${avg_price:,.0f}\n"

 tier_summary = f"\n📋 角色數：{len(g['char_weights'])}（SS/S級：{len(high_tier_chars)}）"

 send_discord(g["discord"],
 f"{emoji} **{name} 市場行情** | {stats['last_updated']} | 共 {stats['count']} 筆\n"
 f"均價：${thresholds['price_avg']:,.0f} | 均金角：{thresholds['gold_char_avg']:.1f}\n"
 f"純角CP：均值 {thresholds['cp1_avg']:.2f} → 門檻 ≤ {thresholds['cp1_threshold']:.2f}\n"
 f"含武CP：均值 {thresholds['cp2_avg']:.2f} → 門檻 ≤ {thresholds['cp2_threshold']:.2f}\n"
 f"加權CP：均值 {thresholds['cpw_avg']:.2f} → 門檻 ≤ {thresholds['cpw_threshold']:.2f}"
 + seller_summary + tier_summary
 )

 print(f"\n🔍 [{name}] 抓現有賣場（{MAX_PAGES} 頁）...")
 listings = scrape_pages(
 main_page, g["list_url"], MAX_PAGES, "賣場",
 do_detail=True, detail_page=detail_page,
 char_weights=g["char_weights"], alias_map=g["alias_map"])

 for r in listings:
 r['estimated_profit'] = estimate_profit(r, stats)

 valid = [r for r in listings if r['cp1'] != float('inf') or r['cpw'] != float('inf')]
 print(f" 有效帳號：{len(valid)} 個")

 today_str = datetime.now().strftime("%Y-%m-%d")
 for r in listings:
 if r['url'] not in listing_seen_map:
 listing_seen_map[r['url']] = today_str
 save_listing_seen(g["listing_seen_file"], listing_seen_map)

 game_tracker = price_tracker.get(name, {})
 game_tracker = check_price_drop(game_tracker, valid, DISCORD_PRICE_DROP, emoji, name)
 price_tracker[name] = game_tracker

 update_excel(g["excel"], valid, thresholds, sellers)
 try:
 ws_google = init_gsheet(gc, name)
 update_gsheet(ws_google, valid, thresholds, sellers)
 except Exception as e:
 print(f" Google Sheets 失敗：{e}")

 listing_seen_set = set(listing_seen_map.keys())
 new_good = [r for r in valid if (
 r['url'] not in listing_seen_set or
 listing_seen_map.get(r['url']) == today_str
 ) and (
 r['cp1'] <= thresholds['cp1_threshold'] or
 r['cp2'] <= thresholds['cp2_threshold'] or
 r['cpw'] <= thresholds['cpw_threshold']
 )]

 print(f" 新上架且優於均值：{len(new_good)} 個")
 if new_good:
 msg = f"{emoji} **{name} 新上架！優於市場均值（共{len(new_good)}個）** | {datetime.now().strftime('%H:%M')}\n{'─'*38}\n"
 for i, r in enumerate(new_good):
 const_tag = f"\n⭐ **滿命：{', '.join(r['max_const'])}**" if r.get('max_const') else ""
 profit = r.get('estimated_profit')
 profit_tag = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
 post_time = r.get('post_time', '')
 time_tag = f"🆕 **{post_time}**" if is_recent(post_time) else f"🕐{post_time}"
 sid = r.get('seller_id', '')
 seller_tag = f" | 🍽️**大盤商{sid}**" if sid and sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD else (f" | 賣家：{sid}" if sid else "")
 line = (
 f"**#{i+1}** {time_tag} | ${r['price']:,}{seller_tag}\n"
 f"{r['title']}{const_tag}{profit_tag}\n"
 f"純角{r['cp1']:.1f} 含武{r['cp2']:.1f} 加權{r['cpw']:.1f}\n"
 f"{r['url']}\n\n"
 )
 if len(msg) + len(line) > 1900:
 send_discord(g["discord"], msg)
 msg = f"{emoji} **（續）**\n"
 msg += line
 if msg.strip():
 send_discord(g["discord"], msg)

 if thresholds['price_avg'] > 0:
 bargain_limit = thresholds['price_avg'] * BARGAIN_THRESHOLD
 bargains = [r for r in valid if r['price'] <= bargain_limit and r['gold_char'] > 0]
 print(f" 撿漏專區：{len(bargains)} 個")
 if bargains:
 msg = f"🎣 **{name} 撿漏專區** | 均價${thresholds['price_avg']:,.0f} → 門檻≤${bargain_limit:,.0f}\n{'─'*38}\n"
 for i, r in enumerate(bargains[:10]):
 profit = r.get('estimated_profit')
 profit_tag = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
 line = (
 f"**#{i+1}** ${r['price']:,} | 金角{r['gold_char']} | 純角CP{r['cp1']:.1f}{profit_tag}\n"
 f"{r['title']}\n{r['url']}\n\n"
 )
 if len(msg) + len(line) > 1900:
 send_discord(g["discord_bargain"], msg)
 msg = f"🎣 **{name} 撿漏（續）**\n"
 msg += line
 if msg.strip():
 send_discord(g["discord_bargain"], msg)

 maxconst_hits = [r for r in valid if any(c in r.get('max_const', []) for c in new_chars)]
 print(f" 新角滿命監控：{len(maxconst_hits)} 個")
 if maxconst_hits:
 msg = f"🌟 **{name} 新角滿命監控** | {datetime.now().strftime('%H:%M')}\n{'─'*38}\n"
 for i, r in enumerate(maxconst_hits):
 hit_chars = [c for c in r.get('max_const', []) if c in new_chars]
 profit = r.get('estimated_profit')
 profit_tag = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
 line = (
 f"**#{i+1}** ⭐ **滿命：{', '.join(hit_chars)}**\n"
 f"${r['price']:,} | {r['title']}{profit_tag}\n"
 f"純角{r['cp1']:.1f} 含武{r['cp2']:.1f}\n{r['url']}\n\n"
 )
 if len(msg) + len(line) > 1900:
 send_discord(g["discord_maxconst"], msg)
 msg = f"🌟 **{name} 新角滿命（續）**\n"
 msg += line
 if msg.strip():
 send_discord(g["discord_maxconst"], msg)

 budget = [r for r in valid if r['price'] <= 2000]
 premium = [r for r in valid if r['price'] > 2000]
 groups = [
 ("💰 平價(≤2000) 純角CP Top3", budget, "cp1", 3),
 ("💰 平價(≤2000) 含武CP Top2", budget, "cp2", 2),
 ("💎 高階(>2000) 純角CP Top3", premium, "cp1", 3),
 ("💎 高階(>2000) 含武CP Top2", premium, "cp2", 2),
 ("💰 平價(≤2000) 加權CP Top3", budget, "cpw", 3),
 ("💎 高階(>2000) 加權CP Top2", premium, "cpw", 2),
 ]
 for group_title, items, cp_key, n in groups:
 sorted_items = sorted(
 [x for x in items if x[cp_key] != float('inf')],
 key=lambda x: x[cp_key])[:n]
 if not sorted_items:
 continue
 msg = f"{emoji} **{name} {group_title}**\n{'─'*38}\n"
 for i, r in enumerate(sorted_items):
 msg += f"**#{i+1}** {format_item(r, cp_key, sellers=sellers)}\n\n"
 send_discord(g["discord"], msg)

def run_scrape():
 print("="*60)
 print(f"🚀 開始執行 Version 1.1 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
 print("="*60)

 GAMES = build_games_config()

 try:
 gc = get_gsheet()
 print("✅ Google Sheets 連線成功")
 except Exception as e:
 print(f"⚠️ Google Sheets 連線失敗：{e}")
 gc = None

 price_tracker = load_price_tracker()

 with sync_playwright() as p:
 browser = p.chromium.launch(headless=True)
 main_page = browser.new_page(
 user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
 detail_page = browser.new_page(
 user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")

 for game_key in ["原神", "鳴潮", "崩鐵", "絕區零"]:
 try:
 run_game(main_page, detail_page, game_key, GAMES[game_key], gc, price_tracker)
 except Exception as e:
 print(f"❌ {game_key} 執行錯誤：{e}")

 browser.close()

 save_price_tracker(price_tracker)
 print(f"\n✅ 全部完成！{datetime.now().strftime('%H:%M:%S')}")

def run_daily_maintenance():
 """每日例行：統一大盤商標記 & 成交天數推算補寫回 Google Sheet。"""
 print("⏰ 執行每日維護 (大盤商標記 & 天數回補)...")
 try:
  import subprocess, sys
  subprocess.run(
   [sys.executable, os.path.join(os.path.dirname(os.path.abspath(__file__)), "daily_maintenance.py")],
   check=True
  )
 except Exception as e:
  print(f"每日維護失敗: {e}")

if __name__ == "__main__":
 print("⏰ 排程啟動，每30分鐘執行一次（立即先跑一次）")
 run_scrape()
 schedule.every(30).minutes.do(run_scrape)
 schedule.every().day.at("18:00").do(run_daily_maintenance)
 while True:
 schedule.run_pending()
 time.sleep(60)
