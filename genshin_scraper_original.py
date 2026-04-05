from playwright.sync_api import sync_playwright
import re, time, random, requests, json, os, schedule
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials

try:
    from generate_chart import generate_trend_chart
except ImportError:
    generate_trend_chart = None
try:
    import pymongo
    HAS_PYMONGO = True
except ImportError:
    HAS_PYMONGO = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

BASE_URL = "https://www.8591.com.tw"
MAX_PAGES = 10
TRASH_KEYWORDS = ["徵", "代練", "初始號"]
BIG_SELLER_THRESHOLD = 5
RECENT_DAYS = 10
PRICE_DROP_THRESHOLD = 0.15
BARGAIN_THRESHOLD = 0.50
PLATFORM_FEE = 0.94
SPREADSHEET_ID = "1SOt-2DwJVEcEgvuvQfAvW6ue6WcrnvywxPbKIJFEcYI"
GCP_KEY_FILE = os.path.join(BASE_DIR, "gcp_key.json")  # fallback 時才用
PRICE_TRACKER_FILE = os.path.join(BASE_DIR, "price_tracker.json")

TIER_LIST_FILES = {
    "原神": os.path.join(BASE_DIR, "genshin_tier_list.json"),
    "鳴潮": os.path.join(BASE_DIR, "wutheringwaves_tier_list.json"),
    "崩鐵": os.path.join(BASE_DIR, "hsr_tier_list.json"),
}

TIER_WEIGHTS = {
    "tierSS": 10,
    "tierS":  7,
    "tierA":  4,
    "tierB":  2,
    "tierC":  1,
    "tierD":  1,
}

HIGH_TIER_LEVELS = {"tierSS", "tierS"}

HEADERS = ["發現時間", "上架時間", "標題", "價格", "金角", "金武/專",
           "純角CP", "含武CP", "加權CP", "預估獲利", "滿命角色", "優於均值", "賣家ID", "連結", "歷來低價", "歷來高價"]

COMPLETED_HEADERS = [
    "成交發現日", "上架時間", "售出所需天數",
    "標題", "價格", "金角", "金武/專",
    "滿命角色", "高Tier角色(SS/S)",
    "純角CP", "含武CP", "加權CP",
    "賣家ID", "連結", "歷來低價", "歷來高價"
]

# ===================== Tier List 載入 =====================

def load_tier_weights(game_name):
    filepath = TIER_LIST_FILES.get(game_name)
    if not filepath or not os.path.exists(filepath):
        print(f"  ⚠️ 找不到 {game_name} Tier List，使用預設權重")
        return {}, [], {}, set()

    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    char_weights = {}
    new_chars = data.get("highValueFor8591", [])
    high_tier_chars = set()

    for tier, base_weight in TIER_WEIGHTS.items():
        for char in data.get(tier, []):
            name = char["name"]
            consensus = char.get("consensus", 1)
            weight = base_weight + min(consensus - 1, 3)
            char_weights[name] = weight
            if tier in HIGH_TIER_LEVELS:
                high_tier_chars.add(name)

    print(f"  📊 {game_name} Tier List 載入：{len(char_weights)} 個角色（SS/S級：{len(high_tier_chars)}個）")
    return char_weights, new_chars, {}, high_tier_chars

def build_games_config():
    base_alias = {
        "原神": {
            "水神": "芙寧娜", "芙芙": "芙寧娜",
            "草神": "納西妲", "小草神": "納西妲",
            "雷神": "雷電將軍", "影": "雷電將軍",
            "萬葉": "楓原萬葉", "葉天帝": "楓原萬葉",
            "水龍": "那維萊特", "水龍王": "那維萊特",
            "火神": "瑪薇卡",
            "僕人": "阿蕾奇諾", "小丑": "阿蕾奇諾",
            "綾華": "神里綾華", "達達": "達達利亞",
            "鍾離": "鐘離",
        },
        "鳴潮": {"卡卡": "卡卡羅"},
        "崩鐵": {"花火": "花火", "阮梅": "阮•梅"},
    }

    games = {
        "原神": {
            "emoji": "⚙️",
            "list_url": "https://www.8591.com.tw/v3/mall/list/34169?searchGame=34169&searchServer=34170&searchType=2&priceStart=100&priceEnd=200000&post_time_sort=1",
            "completed_url": "https://www.8591.com.tw/v3/mall/list/34169?searchGame=34169&searchServer=34170&searchType=2&priceStart=100&priceEnd=200000&completed=1&post_time_sort=1",
            "discord": "https://discord.com/api/webhooks/1483040957397205014/kyBaVtkZ4s0ECMBdNz0RT1eVG8A4irguVG-VzsNBL_TVTd5wTWU4eHHZHpq2cRnMjNXY",
            "discord_bargain": "https://discord.com/api/webhooks/1484537008935276624/VemiDQp698IAADQOFWUnLU9x5kDzlZ2NfwtqdSk9FdIlHcvFnyyAV_LNeMXm9qbgHfzr",
            "discord_maxconst": "https://discord.com/api/webhooks/1484537441405767773/5Jxr2h9BzgkgeVj6adxFA7_Eysx-cOI7PlI-YmfrdTq_ukKob21jLwIc9QRkIsHyGynV",
            "excel": os.path.join(BASE_DIR, "genshin_listings.xlsx"),
            "stats_file": os.path.join(BASE_DIR, "gs_market_stats.json"),
            "history_file": os.path.join(BASE_DIR, "gs_completed_history.json"),
            "listing_seen_file": os.path.join(BASE_DIR, "gs_listing_seen.json"),
            "seller_file": os.path.join(BASE_DIR, "gs_sellers.json"),
        },
        "鳴潮": {
            "emoji": "🌊",
            "list_url": "https://www.8591.com.tw/v3/mall/list/53396?searchGame=53396&searchServer=53397&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&post_time_sort=1",
            "completed_url": "https://www.8591.com.tw/v3/mall/list/53396?searchGame=53396&searchServer=53397&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&completed=1&post_time_sort=1",
            "discord": "https://discord.com/api/webhooks/1483459423434182798/iDMSDYDlZ5bp0_sMPHCFUISyYQxlkO5fzMJP9jo6NLXEHC_AGZkMN8Nb0SmoCuk-c2P9",
            "discord_bargain": "https://discord.com/api/webhooks/1484537092527620117/dCVNGjSHXuTj3MO24vlndbaHuXcqKpolgIMckjpyNsBurXYwEzuwCrIr3LpUI_C0ilUI",
            "discord_maxconst": "https://discord.com/api/webhooks/1484537497827541003/2ZMhGXZeBXq7vmwEUk_fZ6OgtRiUkShi0dQ2ZE2Z8M_XQLK6lrDp356offRRbJB4u94R",
            "excel": os.path.join(BASE_DIR, "wuwa_listings.xlsx"),
            "stats_file": os.path.join(BASE_DIR, "ww_market_stats.json"),
            "history_file": os.path.join(BASE_DIR, "ww_completed_history.json"),
            "listing_seen_file": os.path.join(BASE_DIR, "ww_listing_seen.json"),
            "seller_file": os.path.join(BASE_DIR, "ww_sellers.json"),
        },
        "崩鐵": {
            "emoji": "🚂",
            "list_url": "https://www.8591.com.tw/v3/mall/list/44693?searchGame=44693&searchServer=53160&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&post_time_sort=1",
            "completed_url": "https://www.8591.com.tw/v3/mall/list/44693?searchGame=44693&searchServer=53160&searchType=2&accountTag=3&priceStart=100&priceEnd=200000&completed=1&post_time_sort=2",
            "discord": "https://discord.com/api/webhooks/1483469454376566943/QIaka_rST9Af8dQayNIKf11zM4a6X06k-3MKFIHbB0kK9AJpcR3Lp6fSys_xeA2oGlZT",
            "discord_bargain": "https://discord.com/api/webhooks/1484537146915291176/gHNGq_1m3j_jTJfZzGK2LSOKUxg1QAHmi_ejCP2Frtb9Qg0X1tkVBC4PihH34WObL0u1",
            "discord_maxconst": "https://discord.com/api/webhooks/1484537547689431222/36fxNxtjrLg2LAmvwbrKHE4EiWpor3uJD8mJxQUghpiZd3X1GncMHOJSgG0TH4V32LwP",
            "excel": os.path.join(BASE_DIR, "starrail_listings.xlsx"),
            "stats_file": os.path.join(BASE_DIR, "sr_market_stats.json"),
            "history_file": os.path.join(BASE_DIR, "sr_completed_history.json"),
            "listing_seen_file": os.path.join(BASE_DIR, "sr_listing_seen.json"),
            "seller_file": os.path.join(BASE_DIR, "sr_sellers.json"),
        },
    }

    for game_name, g in games.items():
        char_weights, new_chars, _, high_tier_chars = load_tier_weights(game_name)
        g["char_weights"] = char_weights
        g["new_chars"] = new_chars
        g["alias_map"] = base_alias.get(game_name, {})
        g["high_tier_chars"] = high_tier_chars

    return games

# ===================== listing_seen 新版（含時間戳記與價格區間）=====================

def __migrate_seen_map(seen_map):
    for k, v in list(seen_map.items()):
        if isinstance(v, str):
            seen_map[k] = {"date": v}
    return seen_map

def load_listing_seen(filepath):
    db = get_mongo_db()
    if db is not None:
        doc = db["listing_seen"].find_one({"_id": _mongo_key(filepath)})
        if doc:
            seen_map = doc.get("seen_map", {})
            for url in doc.get("urls", []):
                if url not in seen_map:
                    seen_map[url] = {"date": ""}
            return __migrate_seen_map(seen_map)
        return {}
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        urls = data.get("urls", [])
        seen_map = data.get("seen_map", {})
        for url in urls:
            if url not in seen_map:
                seen_map[url] = {"date": ""}
        return __migrate_seen_map(seen_map)
    return {}

def save_listing_seen(filepath, seen_map):
    db = get_mongo_db()
    if db is not None:
        db["listing_seen"].replace_one(
            {"_id": _mongo_key(filepath)},
            {"_id": _mongo_key(filepath),
             "seen_map": seen_map,
             "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")},
            upsert=True
        )
    else:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({
                "urls": list(seen_map.keys()),
                "seen_map": seen_map,
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")
            }, f, ensure_ascii=False, indent=2)

def calc_days_on_market(post_time_str, seen_map, url, seller_id="", title="", is_big_seller=False):
    """
    計算售出所需天數：
    1. 優先用詳情頁抓到的上架時間
    2. 其次用 listing_seen 的首次發現時間（精確 URL 比對）
    3. 再其次用 title 作精確比對（同一個賣家改價重拋，但標題沒改）
    4. 最後傳入 seller_id 作 fallback（改標題又改價，只能用賣家最早出現時間估算。*排除大盤商*）
    """
    today = datetime.now()

    # 方法A：用詳情頁的上架時間
    if post_time_str and post_time_str != "-":
        for fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
            try:
                d = datetime.strptime(post_time_str.strip(), fmt)
                days = (today - d).days
                return str(days)
            except:
                continue

    # 方法B：用 listing_seen 首次發現時間（精確 URL）
    first_seen = seen_map.get(url, "")
    if isinstance(first_seen, dict):
        first_seen = first_seen.get("date", "")
    if first_seen:
        try:
            d = datetime.strptime(first_seen, "%Y-%m-%d")
            days = (today - d).days
            return f"≥{days}"  # 加 ≥ 表示這是最少天數，實際可能更長
        except:
            pass

    # 方法C：精確標題比對（同一商品改價重拋，標題不變）
    if title:
        title_idx = seen_map.get("__title_idx__", {})
        earliest_by_title = title_idx.get(title, "")
        if earliest_by_title:
            try:
                d = datetime.strptime(earliest_by_title, "%Y-%m-%d")
                days = (today - d).days
                return f"≥{days}*"  # 加 * 表示是用相同標題推算
            except:
                pass

    # 方法D：同一賣家 fallback（賣家改價且改標題，只能用賣家最早活耀日估計。排除大盤商因為他們商品太多）
    if seller_id and not is_big_seller:
        seller_idx = seen_map.get("__seller_idx__", {})
        earliest = seller_idx.get(seller_id, "")
        if earliest:
            try:
                d = datetime.strptime(earliest, "%Y-%m-%d")
                days = (today - d).days
                return f"≥{days}†"  # 加 † 表示是同賣家估算，非精確
            except:
                pass

    return "-"

def apply_price_updates_to_gsheet(ws, updates):
    """將價格範圍異動 (歷來低價/高價) 批次更新回 Google Sheets 的 O/P 欄位"""
    if not ws or not updates:
        return
    try:
        urls_in_sheet = ws.col_values(14)
        updates_batch = []
        for u in updates:
            try:
                row_idx = urls_in_sheet.index(u["url"]) + 1
                updates_batch.append({
                    'range': f'O{row_idx}:P{row_idx}',
                    'values': [[u["min_price"], u["max_price"]]]
                })
            except ValueError:
                continue
        if updates_batch:
            ws.batch_update(updates_batch)
            print(f"  Google Sheets 更新：同步了 {len(updates_batch)} 筆歷史標價極值")
    except Exception as e:
        print(f"  Google Sheets 標價極值更新失敗：{e}")

# ===================== Discord / 警報功能 =====================

_mongo_client = None
_mongo_db = None

def get_mongo_db():
    """
    若環境變數 MONGODB_URI 存在且 pymongo 已安裝，回傳 DB 物件。
    否則回傳 None，讓各函式 fallback 到本機 JSON 檔。
    """
    global _mongo_client, _mongo_db
    if not HAS_PYMONGO:
        return None
    uri = os.environ.get("MONGODB_URI", "")
    if not uri:
        return None
    if _mongo_db is None:
        _mongo_client = pymongo.MongoClient(uri, serverSelectionTimeoutMS=8000)
        _mongo_db = _mongo_client["genshin_scraper"]
        print("  ✅ MongoDB 連線成功")
    return _mongo_db

def _mongo_key(filepath):
    """用檔案名稱（不含副檔名）作為 MongoDB document _id。"""
    return os.path.basename(filepath).replace(".json", "")

# ===================== Google Sheets =====================

def get_gsheet():
    scopes = ["https://spreadsheets.google.com/feeds",
              "https://www.googleapis.com/auth/drive"]
    # 優先從環境變數讀取（Railway 部署用）
    gcp_key_json = os.environ.get("GCP_KEY_JSON", "")
    if gcp_key_json:
        info = json.loads(gcp_key_json)
        creds = Credentials.from_service_account_info(info, scopes=scopes)
    else:
        # fallback：本機開發時讀 gcp_key.json 檔案
        creds = Credentials.from_service_account_file(GCP_KEY_FILE, scopes=scopes)
    return gspread.authorize(creds)


def init_gsheet(gc, game_name):
    try:
        sh = gc.open_by_key(SPREADSHEET_ID)
        try:
            ws = sh.worksheet(game_name)
        except:
            ws = sh.add_worksheet(title=game_name, rows=5000, cols=16)
        if not ws.get_all_values() or ws.cell(1, 1).value != "發現時間":
            ws.insert_row(HEADERS, 1)
        if ws.row_count < 5000 or ws.col_count < 16:
            ws.resize(rows=5000, cols=16)
        return ws
    except Exception as e:
        print(f"  Google Sheets 初始化失敗：{e}")
        return None

def init_gsheet_completed(gc, game_name):
    sheet_name = f"{game_name}-成交紀錄"
    try:
        sh = gc.open_by_key(SPREADSHEET_ID)
        try:
            ws = sh.worksheet(sheet_name)
        except:
            ws = sh.add_worksheet(title=sheet_name, rows=5000, cols=16)
        if not ws.get_all_values() or ws.cell(1, 1).value != "成交發現日":
            ws.insert_row(COMPLETED_HEADERS, 1)
        if ws.row_count < 5000 or ws.col_count < 16:
            ws.resize(rows=5000, cols=16)
        return ws
    except Exception as e:
        print(f"  成交紀錄分頁初始化失敗：{e}")
        return None

def gsheet_insert_with_retry(ws, row, max_retries=3):
    """帶 retry 的單行插入，避免 API 限流導致資料漏記。"""
    for attempt in range(max_retries):
        try:
            ws.insert_row(row, 2)
            time.sleep(2)
            return True
        except Exception as e:
            wait = 15 * (attempt + 1)
            print(f"    ⚠️ insert_row 失敗（第{attempt+1}次）：{e}，等待{wait}s 後重試...")
            time.sleep(wait)
    print(f"    ❌ insert_row 連續失敗 {max_retries} 次，跳過此行")
    return False

def gsheet_batch_insert(ws, rows_to_add):
    """回傳實際成功寫入的 URL 清單（第14欄，index 13）。"""
    if not rows_to_add:
        return []
    written_urls = []
    for i in range(0, len(rows_to_add), 10):
        batch = rows_to_add[i:i+10]
        for row in batch:
            ok = gsheet_insert_with_retry(ws, row)
            if ok:
                written_urls.append(row[13])  # URL 在第14欄
        if i + 10 < len(rows_to_add):
            print(f"    已寫入 {i+len(batch)}/{len(rows_to_add)}...")
            time.sleep(10)
    return written_urls

def gsheet_update_prices(ws, updates):
    """更新已存在的商品價格區間"""
    if not ws or not updates:
        return
    try:
        urls = ws.col_values(14)  # 第 14 欄是網址
        url_idx_map = {url: i+1 for i, url in enumerate(urls)}
        
        batch_data = []
        for upd in updates:
            row_idx = url_idx_map.get(upd["url"])
            if row_idx:
                batch_data.append({
                    'range': f'O{row_idx}:P{row_idx}',
                    'values': [[upd["min_price"], upd["max_price"]]]
                })
        
        if batch_data:
            ws.batch_update(batch_data)
            
            # 加上深綠色粗體格式，標示出曾經改過價格的紀錄
            format_requests = []
            for upd in updates:
                row_idx = url_idx_map.get(upd["url"])
                if row_idx:
                    format_requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": ws.id,
                                "startRowIndex": row_idx - 1,
                                "endRowIndex": row_idx,
                                "startColumnIndex": 14, # 第 15 欄 (O)
                                "endColumnIndex": 16    # 第 16 欄 (P)
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "textFormat": {
                                        # 深綠色 rgb(0, 128, 0)
                                        "foregroundColor": {"red": 0.0, "green": 0.5, "blue": 0.0},
                                        "bold": True
                                    }
                                }
                            },
                            "fields": "userEnteredFormat(textFormat)"
                        }
                    })
            if format_requests:
                ws.spreadsheet.batch_update({"requests": format_requests})
                
            print(f"  Google Sheets 追蹤更新：{len(batch_data)} 筆改價紀錄（已標示深綠色）")
    except Exception as e:
        print(f"  Google Sheets 改價更新失敗：{e}")

def update_gsheet(ws, new_items, thresholds, sellers):
    if not ws:
        return
    try:
        existing = ws.col_values(14)
        existing_urls = set(existing[1:])
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        rows_to_add = []
        for r in new_items:
            if r['url'] in existing_urls:
                continue
            cp1, cp2, cpw = r['cp1'], r['cp2'], r['cpw']
            is_good = (cp1 <= thresholds['cp1_threshold'] or
                       cp2 <= thresholds['cp2_threshold'] or
                       cpw <= thresholds['cpw_threshold'])
            good_str = "✅ 優於均值" if is_good else ""
            const_str = ", ".join(r.get('max_const', []))
            cp1_str = f"{cp1:.2f}" if cp1 != float('inf') else "-"
            cp2_str = f"{cp2:.2f}" if cp2 != float('inf') else "-"
            cpw_str = f"{cpw:.2f}" if cpw != float('inf') else "-"
            profit = r.get('estimated_profit')
            profit_str = f"+${profit:,.0f}" if profit and profit > 0 else (f"-${abs(profit):,.0f}" if profit else "-")
            sid = r.get('seller_id', '')
            is_big = sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD
            seller_str = f"🍽️{sid}" if is_big else sid
            rows_to_add.append([
                now_str, r.get('post_time', ''), r['title'], r['price'],
                r['gold_char'], r['gold_weap'], cp1_str, cp2_str, cpw_str,
                profit_str, const_str, good_str, seller_str, r['url'],
                r['price'], r['price']
            ])
        if not rows_to_add:
            print("  Google Sheets：無新資料")
            return
        written = gsheet_batch_insert(ws, rows_to_add)
        
        # 防止 gspread.insert_row 繼承上一列被標綠色的格式，強制把新插入的 O、P 欄設回黑色正常字體
        if written:
            try:
                ws.spreadsheet.batch_update({"requests": [{
                    "repeatCell": {
                        "range": {
                            "sheetId": ws.id,
                            "startRowIndex": 1, # Row 2 (0-indexed)
                            "endRowIndex": 1 + len(rows_to_add), # 剛插入的數量範圍
                            "startColumnIndex": 14,
                            "endColumnIndex": 16
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "textFormat": {
                                    "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 0.0}, # 黑色
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(textFormat)"
                    }
                }]})
            except Exception as fe:
                print(f"  Google Sheets 格式重置失敗：{fe}")
                
        print(f"  Google Sheets 更新：新增 {len(written)}/{len(rows_to_add)} 筆")
    except Exception as e:
        print(f"  Google Sheets 更新失敗：{e}")

def update_gsheet_completed(ws, new_trades, sellers, seen_map, high_tier_chars):
    if not ws or not new_trades:
        return
    try:
        existing = ws.col_values(14)
        existing_urls = set(existing[1:])
        now_str = datetime.now().strftime("%Y-%m-%d")
        rows_to_add = []

        for r in new_trades:
            if r['url'] in existing_urls:
                continue
            cp1, cp2, cpw = r['cp1'], r['cp2'], r['cpw']
            cp1_str = f"{cp1:.2f}" if cp1 != float('inf') else "-"
            cp2_str = f"{cp2:.2f}" if cp2 != float('inf') else "-"
            cpw_str = f"{cpw:.2f}" if cpw != float('inf') else "-"

            # 上架時間（詳情頁抓到的）
            post_time = r.get('post_time', '')
            # 過濾掉括號數字（列表頁誤讀的成交數）
            if post_time and re.match(r'^\(\d+\)$', post_time.strip()):
                post_time = ''

            # 賣家ID與大盤商判定
            sid = r.get('seller_id', '')
            is_big = sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD
            seller_str = f"🍽️{sid}（大盤商）" if is_big else sid

            # 售出所需天數（優先詳情頁時間 -> listing_seen -> title -> 最後 seller fallback [排除大盤商]）
            days = calc_days_on_market(post_time, seen_map, r['url'], sid, r.get('title', ''), is_big)

            const_str = ", ".join(r.get('max_const', []))

            # 高Tier角色
            title = r.get('title', '')
            high_tier_found = [char for char in high_tier_chars if char in title]
            high_tier_str = ", ".join(high_tier_found) if high_tier_found else "-"

            rows_to_add.append([
                now_str,
                post_time if post_time else "-",
                days,
                r['title'],
                r['price'],
                r['gold_char'],
                r['gold_weap'],
                const_str if const_str else "-",
                high_tier_str,
                cp1_str,
                cp2_str,
                cpw_str,
                seller_str,
                r['url'],
                # 成交紀錄就照搬原本標價當作低高價
                r['price'],
                r['price']
            ])
            # ⚠️ 不在這裡 add(r['url'])，等 batch_insert 確認成功再標記

        if not rows_to_add:
            print("  成交紀錄：無新資料")
            return
        written = gsheet_batch_insert(ws, rows_to_add)
        print(f"  成交紀錄更新：新增 {len(written)}/{len(rows_to_add)} 筆")
        if len(written) < len(rows_to_add):
            failed = [r[13] for r in rows_to_add if r[13] not in set(written)]
            print(f"  ⚠️ 以下 {len(failed)} 筆寫入失敗，下次執行會重試：")
            for url in failed:
                print(f"    {url}")
    except Exception as e:
        print(f"  成交紀錄更新失敗：{e}")

# ===================== 降價追蹤 =====================

def load_price_tracker():
    db = get_mongo_db()
    if db is not None:
        doc = db["price_tracker"].find_one({"_id": "price_tracker"})
        return doc.get("data", {}) if doc else {}
    if os.path.exists(PRICE_TRACKER_FILE):
        with open(PRICE_TRACKER_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_price_tracker(tracker):
    db = get_mongo_db()
    if db is not None:
        db["price_tracker"].replace_one(
            {"_id": "price_tracker"},
            {"_id": "price_tracker", "data": tracker,
             "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")},
            upsert=True
        )
    else:
        with open(PRICE_TRACKER_FILE, "w", encoding="utf-8") as f:
            json.dump(tracker, f, ensure_ascii=False, indent=2)

# ===================== 快速首發監控（不開 Playwright，直打 HTML）=====================

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

# 快速監控：只抓第一頁，比 Playwright 快 10x
_FAST_TRACK_SESSION = None

def _get_fast_session():
    global _FAST_TRACK_SESSION
    if _FAST_TRACK_SESSION is None:
        import requests as _req
        s = _req.Session()
        s.verify = False
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
            "Accept-Language": "zh-TW,zh;q=0.9",
            "Referer": "https://www.8591.com.tw/",
        })
        _FAST_TRACK_SESSION = s
    return _FAST_TRACK_SESSION

def fast_fetch_listings(list_url):
    """用 requests 直接抓第一頁 HTML，回傳解析後的 listing list（每筆含 url/title/price）"""
    if not HAS_BS4:
        return []
    import warnings; warnings.filterwarnings("ignore")
    try:
        sess = _get_fast_session()
        r = sess.get(list_url, timeout=10)
        if r.status_code != 200:
            return []
        soup = BeautifulSoup(r.content, "html.parser")
        listings = []
        # 嘗試抓商品區塊（8591 的 HTML 結構）
        items = soup.select(".list-item-line") or soup.select("[class*='item-box']") or soup.select(".commodity-list li")
        for it in items:
            a = it.select_one("a.fc1, a.show-title, a[href*='/v3/mall/detail/']")
            if not a:
                continue
            href = a.get("href", "")
            if not href.startswith("http"):
                href = "https://www.8591.com.tw" + href
            title = a.get("title", "") or a.get_text(strip=True)
            # 抓價格
            price_tag = it.select_one("[class*='price'], .price, .fc-red")
            price_str = price_tag.get_text(strip=True).replace(",", "").replace("$", "").strip() if price_tag else "0"
            try:
                price = int("".join(filter(str.isdigit, price_str)) or "0")
            except:
                price = 0
            if href and price > 0:
                listings.append({"url": href, "title": title, "price": price})
        return listings
    except Exception as e:
        print(f"  [快速監控] 抓取失敗：{e}")
        return []

def fast_track_scan(GAMES):
    """
    快速首發監控：每 2 分鐘掃一次，只看第一頁新商品。
    若發現未見過的優質商品（cpw >= cp 門檻），立刻發 Discord 警報。
    不寫入 Google Sheets，不跑 Playwright，超輕量。
    """
    if not HAS_BS4:
        return

    import warnings; warnings.filterwarnings("ignore")
    print(f"\n⚡ [快速監控] {datetime.now().strftime('%H:%M:%S')} 掃描中...")

    for game_key, g in GAMES.items():
        listing_seen_file = g["listing_seen_file"]
        stats_file        = g["stats_file"]
        char_weights      = g["char_weights"]
        alias_map         = g.get("alias_map", {})
        discord_url       = g["discord"]
        emoji             = g["emoji"]
        high_tier_chars   = g.get("high_tier_chars", set())

        listing_seen_map = load_listing_seen(listing_seen_file)
        stats            = load_stats(stats_file)
        thresholds       = get_thresholds(stats)
        if thresholds["cpw_threshold"] <= 0:
            continue  # 還沒有足夠歷史資料，跳過

        listings = fast_fetch_listings(g["list_url"])
        if not listings:
            continue

        now_str = datetime.now().strftime("%Y-%m-%d")
        alerts  = []
        new_seen = {}

        for item in listings:
            url   = item["url"]
            title = item["title"]
            price = item["price"]

            # 已見過的跳過（不重複通知）
            if url in listing_seen_map:
                continue

            # 標記為本次第一次見到
            listing_seen_map[url] = {"date": now_str, "min_price": price, "max_price": price}
            new_seen[url] = listing_seen_map[url]

            # 快速計算 CP（只用角色關鍵字解析金角數，無法去詳情頁）
            gold_char = 0
            m = re.search(r'(\d+)金', title)
            if m:
                gold_char = int(m.group(1))

            # 計算加權分數（只用已知關鍵字）
            ws = calc_weighted_score(title, char_weights, alias_map)
            cpw = cp_weighted(price, ws) if ws > 0 else float('inf')

            # 判斷是否達到警報門檻（加倍寬鬆，因為沒有完整資訊）
            if cpw == float('inf') or price <= 0:
                continue
            if cpw <= thresholds["cpw_threshold"] * 1.2:  # 門檻內 120% 都警報
                is_high = any(c in title for c in high_tier_chars)
                star    = "⭐" if is_high else ""
                alerts.append(
                    f"{star}**${price:,}** | CP{cpw:.1f} | {title[:50]}\n{url}"
                )

        # 把新見到的商品記入 listing_seen（這樣下次不重複通知）
        if new_seen:
            save_listing_seen(listing_seen_file, listing_seen_map)

        if alerts:
            msg = (
                f"⚡ **{emoji}{game_key} 快速首發警報** | "
                f"{datetime.now().strftime('%H:%M')} | 門檻 CP≤{thresholds['cpw_threshold']:.1f}\n"
                f"{'─'*38}\n"
            )
            for a in alerts[:10]:
                if len(msg) + len(a) + 2 > 1900:
                    send_discord(discord_url, msg)
                    msg = f"⚡ **{game_key}（續）**\n"
                msg += a + "\n\n"
            if msg.strip():
                send_discord(discord_url, msg)
            print(f"  {game_key}：{len(alerts)} 筆新首發警報已發送")
        else:
            print(f"  {game_key}：{len(listings)} 筆已掃，無新首發")



def check_price_drop(tracker, listings, discord_url, emoji, name):
    dropped = []
    for r in listings:
        url, price = r['url'], r['price']
        
        if url in tracker:
            old_price = tracker[url].get('price', price)
            original_price = tracker[url].get('original_price', old_price)
            drop_count = tracker[url].get('drop_count', 0)
            
            if old_price > price:
                drop_count += 1
                drop_pct = (old_price - price) / old_price
                total_drop_pct = (original_price - price) / original_price if original_price else 0
                
                is_panic = total_drop_pct >= 0.30
                is_freq = drop_count >= 2
                
                if drop_pct >= PRICE_DROP_THRESHOLD or is_panic or is_freq:
                    dropped.append({
                        **r, 'old_price': old_price, 'original_price': original_price,
                        'drop_pct': drop_pct, 'total_drop_pct': total_drop_pct,
                        'drop_count': drop_count, 'is_panic': is_panic, 'is_freq': is_freq
                    })
        else:
            original_price = price
            drop_count = 0
            
        tracker[url] = {
            'price': price, 
            'original_price': original_price,
            'drop_count': drop_count,
            'updated': datetime.now().strftime("%Y-%m-%d %H:%M")
        }
        
    if dropped:
        msg = f"{emoji} **🔥 {name} 降價警告！（共{len(dropped)}個）**\n{'─'*38}\n"
        for r in dropped:
            tags = []
            if r.get('is_panic'): tags.append("🚨 斷頭跳水")
            if r.get('is_freq'): tags.append(f"⚠️ 連降{r['drop_count']}次")
            tag_str = f"【{' '.join(tags)}】" if tags else ""
            
            line = (
                f"{tag_str} **這次下殺 {r['drop_pct']*100:.0f}%！** (總跌 {r['total_drop_pct']*100:.0f}%)\n"
                f"~~${r['original_price']:,}~~ → ~~${r['old_price']:,}~~ → **${r['price']:,}**\n"
                f"{r['title']}\n純角{r['cp1']:.1f} 含武{r['cp2']:.1f}\n{r['url']}\n\n"
            )
            if len(msg) + len(line) > 1900:
                send_discord(discord_url, msg)
                msg = f"{emoji} **🔥 降價警告（續）**\n"
            msg += line
        if msg.strip():
            send_discord(discord_url, msg)
    return tracker

# ===================== 獲利預估 =====================

def estimate_profit(r, stats):
    if not stats.get("records"):
        return None
    price, gold_char = r['price'], r['gold_char']
    if gold_char <= 0:
        return None
    similar = [rec for rec in stats["records"]
               if abs(rec.get("gold_char", 0) - gold_char) <= 3 and rec.get("price", 0) > 0]
    if len(similar) < 3:
        if stats["count"] > 0:
            avg_price = stats["price_sum"] / stats["count"]
            ratio = gold_char / max(stats["gold_char_sum"] / stats["count"], 1)
            estimated_resale = avg_price * ratio
        else:
            return None
    else:
        estimated_resale = sum(rec["price"] for rec in similar) / len(similar)
    return round((estimated_resale * PLATFORM_FEE) - price)

# ===================== 核心解析 =====================

def resolve_alias(name, alias_map):
    return alias_map.get(name, name)

def parse_title_smart(title, char_weights, alias_map):
    gold_char, gold_weap, weighted_score = 0, 0, 0
    max_const_chars = []

    m_char = re.search(r'(\d+)金角', title)
    m_weap = re.search(r'(\d+)金(?:武|專)', title)
    if m_char:
        gold_char = int(m_char.group(1))
    if m_weap:
        gold_weap = int(m_weap.group(1))

    plus_patterns = re.findall(r'(\d+)\+(\d+)([\u4e00-\u9fff]{1,4})', title)
    for n1, n2, char in plus_patterns:
        char = resolve_alias(char, alias_map)
        n1, n2 = int(n1), int(n2)
        if gold_char == 0:
            gold_char += n1
        if gold_weap == 0:
            gold_weap += n2
        weight = char_weights.get(char, 1)
        weighted_score += n1 * weight * 10 + n2 * weight * 5
        if n1 >= 6 and char not in max_const_chars:
            max_const_chars.append(char)

    const_patterns = re.findall(r'(\d+)命([\u4e00-\u9fff]{1,4})', title)
    const_patterns += [(b, a) for a, b in re.findall(r'([\u4e00-\u9fff]{1,4})\s*(\d+)命', title)]
    for n, char in const_patterns:
        char = resolve_alias(char, alias_map)
        n = int(n)
        weight = char_weights.get(char, 1)
        weighted_score += (n + 1) * weight * 10
        if n >= 6 and char not in max_const_chars:
            max_const_chars.append(char)

    for pattern in [
        r'(?:滿命|C6|E6)\s*([\u4e00-\u9fff]{2,4})',
        r'([\u4e00-\u9fff]{2,4})\s*(?:滿命|C6|E6)',
        r'([\u4e00-\u9fff]{2,4})\s*6命',
        r'6命\s*([\u4e00-\u9fff]{2,4})',
    ]:
        for c in re.findall(pattern, title):
            c = resolve_alias(c, alias_map)
            if len(c) >= 2 and c not in max_const_chars:
                max_const_chars.append(c)

    if not plus_patterns and not const_patterns:
        for char, weight in char_weights.items():
            aliases = [k for k, v in alias_map.items() if v == char]
            if char in title or any(a in title for a in aliases):
                weighted_score += weight * 10

    return gold_char, gold_weap, weighted_score, max_const_chars

def parse_detail_for_gold(page, url, title):
    try:
        page.goto(url, timeout=30000)
        time.sleep(random.uniform(3, 4))
        # 等待刊登時間出現
        try:
            page.wait_for_selector("text=刊登時間", timeout=5000)
        except:
            pass
        body_text = page.inner_text("body")
        gold_char, gold_weap = 0, 0
        m_char = re.search(r'(\d+)金角', body_text)
        m_weap = re.search(r'(\d+)金(?:武|專)', body_text)
        if m_char:
            gold_char = int(m_char.group(1))
        if m_weap:
            gold_weap = int(m_weap.group(1))
        # 抓上架時間（格式：YYYY-MM-DD 或 YYYY/MM/DD）
        time_matches = re.findall(r'(\d{4}[-/]\d{2}[-/]\d{2})', body_text)
        post_time = time_matches[0] if time_matches else ""
        return gold_char, gold_weap, post_time
    except:
        return 0, 0, ""

def cp_char_only(price, gold_char):
    if price <= 0 or gold_char <= 0:
        return float('inf')
    return price / (gold_char * 10)

def cp_with_weap(price, gold_char, gold_weap):
    total = gold_char * 10 + gold_weap * 5
    if price <= 0 or total <= 0:
        return float('inf')
    return price / total

def cp_weighted(price, weighted_score):
    if price <= 0 or weighted_score <= 0:
        return float('inf')
    return price / weighted_score

def is_recent(post_time_str, days=RECENT_DAYS):
    if not post_time_str:
        return False
    if "分鐘前" in post_time_str or "小時前" in post_time_str:
        return True
    if "天前" in post_time_str:
        m = re.search(r'(\d+)天前', post_time_str)
        if m and int(m.group(1)) <= days:
            return True
    try:
        for fmt in ["%Y-%m-%d", "%Y/%m/%d"]:
            try:
                d = datetime.strptime(post_time_str.strip(), fmt)
                return (datetime.now() - d).days <= days
            except:
                continue
    except:
        pass
    return False

def send_discord(webhook_url, content, image_path=None):
    webhook_url = webhook_url.strip().rstrip('ㄛ').rstrip('/')
    try:
        if image_path and os.path.exists(image_path):
            with open(image_path, "rb") as f:
                r = requests.post(webhook_url, data={"content": content}, files={"file": f})
        else:
            r = requests.post(webhook_url, json={"content": content})
        print(f"  Discord：{r.status_code}")
        time.sleep(0.8)
    except Exception as e:
        print(f"  Discord 失敗：{e}")

def load_stats(filepath):
    db = get_mongo_db()
    if db is not None:
        doc = db["market_stats"].find_one({"_id": _mongo_key(filepath)})
        if doc:
            doc.pop("_id", None)
            if "records" not in doc:
                doc["records"] = []
            return doc
        return {"count": 0, "cp1_sum": 0, "cp2_sum": 0, "cpw_sum": 0,
                "price_sum": 0, "gold_char_sum": 0, "records": [], "last_updated": ""}
    # fallback: 本機 JSON
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        if "records" not in data:
            data["records"] = []
        return data
    return {"count": 0, "cp1_sum": 0, "cp2_sum": 0, "cpw_sum": 0,
            "price_sum": 0, "gold_char_sum": 0, "records": [], "last_updated": ""}

def update_stats(stats, new_trades, filepath):
    valid = [r for r in new_trades if r['cp1'] != float('inf')]
    if "records" not in stats:
        stats["records"] = []
    for r in valid:
        stats["count"] += 1
        stats["cp1_sum"] += r['cp1']
        stats["cp2_sum"] += r['cp2'] if r['cp2'] != float('inf') else r['cp1']
        stats["cpw_sum"] += r['cpw'] if r['cpw'] != float('inf') else r['cp1']
        stats["price_sum"] += r['price']
        stats["gold_char_sum"] += r['gold_char']
        stats["records"].append({
            "date": datetime.now().strftime("%Y-%m-%d"),
            "price": r['price'],
            "gold_char": r['gold_char'],
            "gold_weap": r['gold_weap'],
            "cp1": round(r['cp1'], 2),
            "cp2": round(r['cp2'], 2) if r['cp2'] != float('inf') else None,
            "cpw": round(r['cpw'], 2) if r['cpw'] != float('inf') else None,
        })
    # 不讓 records 無限展開。MongoDB 單文件限制 16MB，保留最近 2000 筆就夠
    stats["records"] = stats["records"][-2000:]
    stats["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    db = get_mongo_db()
    if db is not None:
        db["market_stats"].replace_one(
            {"_id": _mongo_key(filepath)},
            {"_id": _mongo_key(filepath), **stats},
            upsert=True
        )
    else:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)
    return stats

def get_thresholds(stats):
    records = stats.get("records", [])
    if not records:
        return {"cp1_threshold": 30, "cp2_threshold": 25, "cpw_threshold": 20,
                "cp1_avg": 30, "cp2_avg": 25, "cpw_avg": 20,
                "price_avg": 0, "gold_char_avg": 0}

    def get_trimmed_mean(vals):
        if not vals: return 0
        if len(vals) < 5: return sum(vals) / len(vals)
        vals.sort()
        # 剔除最高與最低 15%（解決極端值/亂填資料）
        trim_idx = max(1, int(len(vals) * 0.15))
        trimmed = vals[trim_idx:-trim_idx]
        return sum(trimmed) / len(trimmed) if trimmed else (sum(vals) / len(vals))

    cp1_vals = [r["cp1"] for r in records if r.get("cp1")]
    cp2_vals = [r["cp2"] for r in records if r.get("cp2")]
    cpw_vals = [r["cpw"] for r in records if r.get("cpw")]
    prices = [r["price"] for r in records if r.get("price")]
    golds = [r["gold_char"] for r in records if r.get("gold_char")]

    cp1_avg = get_trimmed_mean(cp1_vals) or 30
    cp2_avg = get_trimmed_mean(cp2_vals) or 25
    cpw_avg = get_trimmed_mean(cpw_vals) or 20

    return {
        "cp1_avg": cp1_avg, "cp2_avg": cp2_avg, "cpw_avg": cpw_avg,
        "cp1_threshold": cp1_avg * 0.8,
        "cp2_threshold": cp2_avg * 0.8,
        "cpw_threshold": cpw_avg * 0.8,
        "price_avg": get_trimmed_mean(prices),
        "gold_char_avg": get_trimmed_mean(golds),
    }

def load_seen(filepath, key="urls"):
    db = get_mongo_db()
    if db is not None:
        doc = db["completed_seen"].find_one({"_id": _mongo_key(filepath)})
        return set(doc.get(key, [])) if doc else set()
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return set(json.load(f).get(key, []))
    return set()

def save_seen(filepath, seen, key="urls"):
    db = get_mongo_db()
    if db is not None:
        db["completed_seen"].replace_one(
            {"_id": _mongo_key(filepath)},
            {"_id": _mongo_key(filepath), key: list(seen),
             "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")},
            upsert=True
        )
    else:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump({key: list(seen),
                       "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")}, f)

def load_sellers(filepath):
    db = get_mongo_db()
    if db is not None:
        doc = db["sellers"].find_one({"_id": _mongo_key(filepath)})
        return doc.get("data", {}) if doc else {}
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def update_sellers(sellers, trades, filepath):
    for r in trades:
        sid = r.get("seller_id", "")
        if not sid:
            continue
        if sid not in sellers:
            sellers[sid] = {"count": 0, "prices": [], "titles": []}
        sellers[sid]["count"] += 1
        sellers[sid]["prices"].append(r["price"])
        sellers[sid]["titles"].append(r["title"][:30])
        sellers[sid]["prices"] = sellers[sid]["prices"][-50:]
        sellers[sid]["titles"] = sellers[sid]["titles"][-50:]
    db = get_mongo_db()
    if db is not None:
        db["sellers"].replace_one(
            {"_id": _mongo_key(filepath)},
            {"_id": _mongo_key(filepath), "data": sellers,
             "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M")},
            upsert=True
        )
    else:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(sellers, f, ensure_ascii=False, indent=2)
    return sellers

def init_excel(filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "賣場列表"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    col_widths = [16, 12, 60, 10, 8, 8, 10, 10, 10, 12, 20, 10, 12, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(filepath)

def update_excel(filepath, new_items, thresholds, sellers):
    if not os.path.exists(filepath):
        init_excel(filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    existing_urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[13]:
            existing_urls.add(row[13])
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    added = 0
    for r in new_items:
        if r['url'] in existing_urls:
            continue
        cp1, cp2, cpw = r['cp1'], r['cp2'], r['cpw']
        is_good = (cp1 <= thresholds['cp1_threshold'] or
                   cp2 <= thresholds['cp2_threshold'] or
                   cpw <= thresholds['cpw_threshold'])
        good_str = "✅ 優於均值" if is_good else ""
        const_str = ", ".join(r.get('max_const', []))
        cp1_str = f"{cp1:.2f}" if cp1 != float('inf') else "-"
        cp2_str = f"{cp2:.2f}" if cp2 != float('inf') else "-"
        cpw_str = f"{cpw:.2f}" if cpw != float('inf') else "-"
        profit = r.get('estimated_profit')
        profit_str = f"+${profit:,.0f}" if profit and profit > 0 else (f"-${abs(profit):,.0f}" if profit else "-")
        sid = r.get('seller_id', '')
        is_big = sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD
        seller_str = f"🍽️{sid}" if is_big else sid
        ws.insert_rows(2)
        row_data = [now_str, r.get('post_time', ''), r['title'], r['price'],
                    r['gold_char'], r['gold_weap'], cp1_str, cp2_str, cpw_str,
                    profit_str, const_str, good_str, seller_str, r['url']]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if is_good and const_str:
                cell.fill = PatternFill("solid", fgColor="FFC000")
            elif is_good:
                cell.fill = PatternFill("solid", fgColor="E2EFDA")
            elif const_str:
                cell.fill = PatternFill("solid", fgColor="FFE699")
            elif is_big:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
        ws.row_dimensions[2].height = 40
        existing_urls.add(r['url'])
        added += 1
    wb.save(filepath)
    print(f"  Excel 更新：新增 {added} 筆 → {filepath}")

def get_item_url(item, title_el):
    href = title_el.get_attribute("href")
    if href and href.startswith("/"):
        return BASE_URL + href
    elif href:
        return href
    else:
        im_el = item.query_selector("a[href^='im://']")
        if im_el:
            im_href = im_el.get_attribute("href") or ""
            m = re.search(r'i=(\d+)', im_href)
            if m:
                return f"{BASE_URL}/v3/mall/detail/{m.group(1)}"
    return ""

def scrape_pages(main_page, base_url, max_pages, label="",
                 stop_at_seen=None, do_detail=False, detail_page=None,
                 char_weights=None, alias_map=None):
    results = []
    price_updates = []
    seen_in_run = set()
    stop_at_seen = stop_at_seen if stop_at_seen is not None else set()

    # 連續遇到已見過的筆數計數（避免全新頁面也停止）
    # 8591 completed list 不是嚴格時間排序，不能靠 hit_old 邏輯直接 break
    # 改為：整頁都是舊的才停，否則繼續掃
    for page_num in range(1, max_pages + 1):
        first_row = (page_num - 1) * 40
        url = f"{base_url}&firstRow={first_row}"
        print(f"  [{label}] 第{page_num}頁...")
        try:
            main_page.goto(url, timeout=60000)
            time.sleep(random.uniform(3, 4))
            items = main_page.query_selector_all("div.list-item")
            if not items:
                items = main_page.query_selector_all(".commodity-list > li")
            if not items:
                print("    無商品，停止。")
                break
            print(f"    找到 {len(items)} 個")

            new_in_page = 0
            
            for item in items:
                try:
                    title_el = item.query_selector("a.show-title")
                    if not title_el:
                        title_el = item.query_selector("span.show-title")
                    
                    seller_el = item.query_selector("a[href^='im://']")
                    seller_id = seller_el.get_attribute("data-fuid") if seller_el else ""

                    if not title_el:
                        continue
                    
                    title = title_el.inner_text().strip()
                    detail_url = get_item_url(item, title_el)
                    
                    price = 0
                    price_el = item.query_selector("span.orange")
                    if not price_el:
                        price_el = item.query_selector("div.list-item-price")
                        
                    if price_el:
                        price_text = re.sub(r'[^\d]', '', price_el.inner_text())
                        if price_text:
                            price = int(price_text)
                    elif seller_el:
                        im_href = seller_el.get_attribute("href") or ""
                        pm = re.search(r'price=(\d+)', im_href)
                        if pm:
                            price = int(pm.group(1))

                    if price <= 0:
                        continue

                    time_el = item.query_selector(".list-item-bread span.ml15")
                    if not time_el:
                        time_el = item.query_selector(".fc3")
                        
                    post_time = time_el.inner_text().strip() if time_el else ""

                    # 過濾列表頁誤讀的問與答數量（例如 (5) 或 (99+)）
                    if post_time and re.match(r'^\(.*\)$', post_time.strip()):
                        post_time = ''

                    if any(kw in title for kw in TRASH_KEYWORDS):
                        continue
                    if price < 100:
                        continue
                    if not detail_url:
                        continue
                    if detail_url in seen_in_run:
                        continue
                    # ★ 關鍵修正：遇到已記錄的 URL 只跳過這一筆，不 break 整個迴圈
                    # 因為 8591 completed 列表排序不嚴格，舊交易可能夾在新交易中間
                    if detail_url in stop_at_seen:
                        seen_in_run.add(detail_url)  # 避免重複判斷
                        # 檢查價格變化 (stop_at_seen 是 dict 的時候)
                        if isinstance(stop_at_seen, dict):
                            val = stop_at_seen[detail_url]
                            if isinstance(val, dict):
                                old_min = val.get("min_price", price)
                                old_max = val.get("max_price", price)
                                if price < old_min:
                                    val["min_price"] = price
                                    price_updates.append({"url": detail_url, "min_price": price, "max_price": old_max})
                                elif price > old_max:
                                    val["max_price"] = price
                                    price_updates.append({"url": detail_url, "min_price": old_min, "max_price": price})
                        continue

                    seen_in_run.add(detail_url)
                    new_in_page += 1
                    gold_char, gold_weap, weighted, max_const = parse_title_smart(
                        title, char_weights, alias_map)

                    if do_detail and detail_page:
                        d_char, d_weap, d_time = parse_detail_for_gold(detail_page, detail_url, title)
                        if d_char and d_char > gold_char:
                            gold_char = d_char
                        if d_weap and d_weap > gold_weap:
                            gold_weap = d_weap
                        if d_time:
                            post_time = d_time  # 詳情頁時間優先

                    results.append({
                        "title": title, "price": price,
                        "gold_char": gold_char, "gold_weap": gold_weap,
                        "weighted": weighted,
                        "cp1": cp_char_only(price, gold_char),
                        "cp2": cp_with_weap(price, gold_char, gold_weap),
                        "cpw": cp_weighted(price, weighted),
                        "max_const": max_const,
                        "post_time": post_time,
                        "seller_id": seller_id,
                        "url": detail_url
                    })
                except Exception as e:
                    continue

            # 整頁都是舊的（且 stop_at_seen 非空，表示確實有在追蹤）
            # → 表示已到達已記錄的區域，後面幾乎不會有新資料，停止
            if stop_at_seen is not None and not isinstance(stop_at_seen, dict) and new_in_page == 0:
                print(f"    整頁皆為已記錄，停止繼續翻頁。")
                break

        except Exception as e:
            print(f"  頁面錯誤：{e}")
            break
        time.sleep(random.uniform(2, 3))
    return results, price_updates

def format_item(r, cp_type="cp1", sellers=None):
    cp_val = r.get(cp_type, float('inf'))
    cp_label = {"cp1": "純角CP", "cp2": "含武CP", "cpw": "加權CP"}[cp_type]
    post_time = r.get('post_time', '')
    time_str = f" | 🆕 **{post_time}**" if is_recent(post_time) else (f" | 🕐 {post_time}" if post_time else "")
    const_str = f"\n⭐ **滿命：{', '.join(r['max_const'])}**" if r.get('max_const') else ""
    profit = r.get('estimated_profit')
    profit_str = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
    seller_id = r.get('seller_id', '')
    if seller_id and sellers and sellers.get(seller_id, {}).get("count", 0) >= BIG_SELLER_THRESHOLD:
        cnt = sellers[seller_id]["count"]
        avg_p = sum(sellers[seller_id]["prices"]) / len(sellers[seller_id]["prices"])
        seller_tag = f" | 🍽️ **大盤商 No.{seller_id}**（售{cnt}次 均${avg_p:,.0f}）"
    elif seller_id:
        seller_tag = f" | 賣家：{seller_id}"
    else:
        seller_tag = ""
    return (
        f"**標題：** {r['title']}\n"
        f"**價格：** ${r['price']:,} | 金角：{r['gold_char']} | 金武/專：{r['gold_weap']}{time_str}{seller_tag}{const_str}{profit_str}\n"
        f"**{cp_label}：** {cp_val:.2f}\n"
        f"**連結：** {r['url']}"
    )

def run_game(main_page, detail_page, game_key, g, gc, price_tracker):
    emoji = g["emoji"]
    name = game_key
    new_chars = g.get("new_chars", [])
    high_tier_chars = g.get("high_tier_chars", set())

    print(f"\n{'='*60}")
    print(f"{emoji} {name} | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  角色權重數：{len(g['char_weights'])} 個（SS/S級：{len(high_tier_chars)}個）")
    print(f"{'='*60}")

    stats = load_stats(g["stats_file"])
    seen_urls = load_seen(g["history_file"], key="seen_urls")
    listing_seen_map = load_listing_seen(g["listing_seen_file"])
    sellers = load_sellers(g["seller_file"])

    print(f"\n🔍 [{name}] 抓歷史成交...")
    new_completed, _ = scrape_pages(
        main_page, g["completed_url"], 100, "已完成",
        stop_at_seen=seen_urls, do_detail=False, detail_page=None,
        char_weights=g["char_weights"], alias_map=g["alias_map"])
    print(f"  新增 {len(new_completed)} 筆")

    if new_completed:
        stats = update_stats(stats, new_completed, g["stats_file"])
        sellers = update_sellers(sellers, new_completed, g["seller_file"])
        for r in new_completed:
            seen_urls.add(r["url"])
        save_seen(g["history_file"], seen_urls, key="seen_urls")

        try:
            ws_completed = init_gsheet_completed(gc, name)
            update_gsheet_completed(ws_completed, new_completed, sellers,
                                     listing_seen_map, high_tier_chars)
        except Exception as e:
            print(f"  成交紀錄寫入失敗：{e}")

    thresholds = get_thresholds(stats)
    print(f"  純角CP：{thresholds['cp1_avg']:.2f} → 門檻 {thresholds['cp1_threshold']:.2f}")
    print(f"  含武CP：{thresholds['cp2_avg']:.2f} → 門檻 {thresholds['cp2_threshold']:.2f}")
    print(f"  加權CP：{thresholds['cpw_avg']:.2f} → 門檻 {thresholds['cpw_threshold']:.2f}")

    big_sellers = {sid: info for sid, info in sellers.items() if info["count"] >= BIG_SELLER_THRESHOLD}
    seller_summary = ""
    if big_sellers:
        top = sorted(big_sellers.items(), key=lambda x: x[1]["count"], reverse=True)[:5]
        seller_summary = "\n🍽️ **常見大盤商：**\n"
        for sid, info in top:
            avg_price = sum(info["prices"]) / len(info["prices"]) if info["prices"] else 0
            seller_summary += f"No.{sid} | 出售{info['count']}次 | 均價${avg_price:,.0f}\n"

    tier_summary = f"\n📋 角色數：{len(g['char_weights'])}（SS/S級：{len(high_tier_chars)}）"

    send_discord(g["discord"],
        f"{emoji} **{name} 市場行情** | {stats['last_updated']} | 共 {stats['count']} 筆\n"
        f"均價：${thresholds['price_avg']:,.0f} | 均金角：{thresholds['gold_char_avg']:.1f}\n"
        f"純角CP：均值 {thresholds['cp1_avg']:.2f} → 門檻 ≤ {thresholds['cp1_threshold']:.2f}\n"
        f"含武CP：均值 {thresholds['cp2_avg']:.2f} → 門檻 ≤ {thresholds['cp2_threshold']:.2f}\n"
        f"加權CP：均值 {thresholds['cpw_avg']:.2f} → 門檻 ≤ {thresholds['cpw_threshold']:.2f}"
        + seller_summary + tier_summary
    )

    print(f"\n🔍 [{name}] 抓現有賣場（{MAX_PAGES} 頁）...")
    listings, active_price_updates = scrape_pages(
        main_page, g["list_url"], MAX_PAGES, "賣場",
        stop_at_seen=listing_seen_map, do_detail=True, detail_page=detail_page,
        char_weights=g["char_weights"], alias_map=g["alias_map"])

    for r in listings:
        r['estimated_profit'] = estimate_profit(r, stats)

    valid = [r for r in listings if r['cp1'] != float('inf') or r['cpw'] != float('inf')]
    print(f"  有效帳號：{len(valid)} 個")

    today_str = datetime.now().strftime("%Y-%m-%d")
    seller_idx = listing_seen_map.setdefault("__seller_idx__", {})
    title_idx = listing_seen_map.setdefault("__title_idx__", {})
    
    for r in listings:
        if r['url'] not in listing_seen_map:
            listing_seen_map[r['url']] = {
                "date": today_str,
                "min_price": r['price'],
                "max_price": r['price']
            }
            
        # 更新賣家索引：僅全變早（賣家首次上架的日期）
        sid = r.get('seller_id', '')
        if sid:
            existing_s = seller_idx.get(sid, "")
            if not existing_s or today_str < existing_s:
                seller_idx[sid] = today_str
                
        # 更新標題索引：紀錄該標題第一次出現的日期
        title_str = r.get('title', '').strip()
        if title_str:
            existing_t = title_idx.get(title_str, "")
            if not existing_t or today_str < existing_t:
                title_idx[title_str] = today_str
                
    save_listing_seen(g["listing_seen_file"], listing_seen_map)

    game_tracker = price_tracker.get(name, {})
    game_tracker = check_price_drop(game_tracker, valid, g["discord"], emoji, name)
    price_tracker[name] = game_tracker

    update_excel(g["excel"], valid, thresholds, sellers)
    try:
        ws_google = init_gsheet(gc, name)
        update_gsheet(ws_google, valid, thresholds, sellers)
        if active_price_updates:
            gsheet_update_prices(ws_google, active_price_updates)
    except Exception as e:
        print(f"  Google Sheets 失敗：{e}")

    listing_seen_set = set(listing_seen_map.keys())
    new_good = [r for r in valid if (
        r['url'] not in listing_seen_set or
        (isinstance(listing_seen_map.get(r['url']), dict) and listing_seen_map.get(r['url']).get('date') == today_str) or
        listing_seen_map.get(r['url']) == today_str
    ) and (
        r['cp1'] <= thresholds['cp1_threshold'] or
        r['cp2'] <= thresholds['cp2_threshold'] or
        r['cpw'] <= thresholds['cpw_threshold']
    )]

    print(f"  新上架且優於均值：{len(new_good)} 個")
    if new_good:
        msg = f"{emoji} **{name} 新上架！優於市場均值（共{len(new_good)}個）** | {datetime.now().strftime('%H:%M')}\n{'─'*38}\n"
        for i, r in enumerate(new_good):
            const_tag = f"\n⭐ **滿命：{', '.join(r['max_const'])}**" if r.get('max_const') else ""
            profit = r.get('estimated_profit')
            profit_tag = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
            post_time = r.get('post_time', '')
            time_tag = f"🆕 **{post_time}**" if is_recent(post_time) else f"🕐{post_time}"
            sid = r.get('seller_id', '')
            seller_tag = f" | 🍽️**大盤商{sid}**" if sid and sellers.get(sid, {}).get("count", 0) >= BIG_SELLER_THRESHOLD else (f" | 賣家：{sid}" if sid else "")
            line = (
                f"**#{i+1}** {time_tag} | ${r['price']:,}{seller_tag}\n"
                f"{r['title']}{const_tag}{profit_tag}\n"
                f"純角{r['cp1']:.1f} 含武{r['cp2']:.1f} 加權{r['cpw']:.1f}\n"
                f"{r['url']}\n\n"
            )
            if len(msg) + len(line) > 1900:
                send_discord(g["discord"], msg)
                msg = f"{emoji} **（續）**\n"
            msg += line
        if msg.strip():
            send_discord(g["discord"], msg)

    if thresholds['price_avg'] > 0:
        bargain_limit = thresholds['price_avg'] * BARGAIN_THRESHOLD
        bargains = [r for r in valid if r['price'] <= bargain_limit and r['gold_char'] > 0]
        print(f"  撿漏專區：{len(bargains)} 個")
        if bargains:
            msg = f"🎣 **{name} 撿漏專區** | 均價${thresholds['price_avg']:,.0f} → 門檻≤${bargain_limit:,.0f}\n{'─'*38}\n"
            for i, r in enumerate(bargains[:10]):
                profit = r.get('estimated_profit')
                profit_tag = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
                line = (
                    f"**#{i+1}** ${r['price']:,} | 金角{r['gold_char']} | 純角CP{r['cp1']:.1f}{profit_tag}\n"
                    f"{r['title']}\n{r['url']}\n\n"
                )
                if len(msg) + len(line) > 1900:
                    send_discord(g["discord_bargain"], msg)
                    msg = f"🎣 **{name} 撿漏（續）**\n"
                msg += line
            if msg.strip():
                send_discord(g["discord_bargain"], msg)

    maxconst_hits = [r for r in valid if any(c in r.get('max_const', []) for c in new_chars)]
    print(f"  新角滿命監控：{len(maxconst_hits)} 個")
    if maxconst_hits:
        msg = f"🌟 **{name} 新角滿命監控** | {datetime.now().strftime('%H:%M')}\n{'─'*38}\n"
        for i, r in enumerate(maxconst_hits):
            hit_chars = [c for c in r.get('max_const', []) if c in new_chars]
            profit = r.get('estimated_profit')
            profit_tag = f"\n💰 **預估獲利：+${profit:,.0f}**" if profit and profit > 0 else ""
            line = (
                f"**#{i+1}** ⭐ **滿命：{', '.join(hit_chars)}**\n"
                f"${r['price']:,} | {r['title']}{profit_tag}\n"
                f"純角{r['cp1']:.1f} 含武{r['cp2']:.1f}\n{r['url']}\n\n"
            )
            if len(msg) + len(line) > 1900:
                send_discord(g["discord_maxconst"], msg)
                msg = f"🌟 **{name} 新角滿命（續）**\n"
            msg += line
        if msg.strip():
            send_discord(g["discord_maxconst"], msg)

    budget = [r for r in valid if r['price'] <= 2000]
    premium = [r for r in valid if r['price'] > 2000]
    groups = [
        ("💰 平價(≤2000) 純角CP Top3",  budget,  "cp1", 3),
        ("💰 平價(≤2000) 含武CP Top2",  budget,  "cp2", 2),
        ("💎 高階(>2000) 純角CP Top3",  premium, "cp1", 3),
        ("💎 高階(>2000) 含武CP Top2",  premium, "cp2", 2),
        ("💰 平價(≤2000) 加權CP Top3",  budget,  "cpw", 3),
        ("💎 高階(>2000) 加權CP Top2",  premium, "cpw", 2),
    ]
    for group_title, items, cp_key, n in groups:
        sorted_items = sorted(
            [x for x in items if x[cp_key] != float('inf')],
            key=lambda x: x[cp_key])[:n]
        if not sorted_items:
            continue
        msg = f"{emoji} **{name} {group_title}**\n{'─'*38}\n"
        for i, r in enumerate(sorted_items):
            msg += f"**#{i+1}** {format_item(r, cp_key, sellers=sellers)}\n\n"
        send_discord(g["discord"], msg)

def run_scrape():
    print("="*60)
    print(f"🚀 開始執行 Version 1.1 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("="*60)

    GAMES = build_games_config()

    try:
        gc = get_gsheet()
        print("✅ Google Sheets 連線成功")
    except Exception as e:
        print(f"⚠️ Google Sheets 連線失敗：{e}")
        gc = None

    price_tracker = load_price_tracker()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        main_page = browser.new_page(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        detail_page = browser.new_page(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")

        for game_key in ["原神", "鳴潮", "崩鐵"]:
            try:
                run_game(main_page, detail_page, game_key, GAMES[game_key], gc, price_tracker)
            except Exception as e:
                print(f"❌ {game_key} 執行錯誤：{e}")

        browser.close()

    save_price_tracker(price_tracker)
    print(f"\n✅ 全部完成！{datetime.now().strftime('%H:%M:%S')}")

def run_trend_charts(GAMES):
    print("\n📊 產製並發送市場趨勢週報...")
    for game_key, g in GAMES.items():
        stats_file = g["stats_file"]
        stats = load_stats(stats_file)
        if generate_trend_chart:
            chart_path = f"trend_{game_key}.png"
            try:
                out = generate_trend_chart(game_key, stats, output_path=chart_path)
                if out and g["discord"]:
                    msg = f"📈 **{g['emoji']} {game_key} 市場動盪趨勢週報**\n為您呈上近30天的日均價與純角CP走勢分析！"
                    send_discord(g["discord"], msg, image_path=out)
            except Exception as e:
                print(f"  ❌ {game_key} 趨勢圖生成失敗：{e}")

if __name__ == "__main__":
    print("⏰ 排程啟動，每30分鐘執行一次（立即先跑一次）")
    print("⚡ 快速監控：每2分鐘掃一次首頁新上架（不用 Playwright，超輕量）")
    GAMES = build_games_config()
    
    # 首次啟動：為展示新功能，強制發送一次趨勢週報
    run_trend_charts(GAMES)
    
    run_scrape()
    schedule.every(30).minutes.do(run_scrape)
    schedule.every(2).minutes.do(lambda: fast_track_scan(GAMES))
    schedule.every().sunday.at("20:00").do(lambda: run_trend_charts(GAMES))
    
    while True:
        schedule.run_pending()
        time.sleep(30)  # 每 30 秒檢查一次排程（支援 2 分鐘精確度）
